from django.shortcuts import render, redirect
from django.http import JsonResponse
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status
from .models import *  
from .serializers import StudentSerializer 
from django.utils import timezone
from datetime import timedelta,time,datetime
import hashlib
import json
from django.core.serializers.json import DjangoJSONEncoder
from django.shortcuts import render, get_object_or_404,redirect
from rest_framework.permissions import IsAuthenticated, AllowAny

from django.contrib.auth import authenticate
from rest_framework.authtoken.models import Token
from django.contrib.auth.decorators import login_required
from .forms import StudentForm
from django.core.paginator import Paginator
from django.db.models import Q # للبحث
from django.contrib import messages
from django.urls import reverse
from django.http import HttpResponse
import csv  
import io
import openpyxl
from openpyxl.styles import Protection, Font, PatternFill
from openpyxl.utils import get_column_letter
import datetime
from datetime import datetime
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.lib.utils import ImageReader
import qrcode
from django.db.models.functions import Trim, Lower
from django.views.decorators.http import require_POST

from PIL import Image, ImageDraw, ImageFont
import zipfile

from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_exempt
from django.conf import settings
import openpyxl
from django.http import HttpResponse
from openpyxl.styles import Font, Fill, PatternFill, Alignment, Protection, Border, Side
import os
from django.contrib.auth import update_session_auth_hash
from django.utils.dateparse import parse_date
import re
from django.utils.timezone import now
from django.db.models import Count
from django.db.models import Prefetch
from django.db.models import Count, Q, Max
import traceback
import urllib.parse
from django.db.models import Sum





# تأكد أن Imports كاينين الفوق
from .models import UserProfile # 👈 ضروري

@api_view(['POST'])
@permission_classes([AllowAny])
def custom_login(request):
    username = request.data.get('username')
    password = request.data.get('password')

    # تنظيف اليوزرنيم من أي فراغات زايدة تقدر تجي من التليفون أو الكيبورد
    if username:
        username = username.strip()

    print(f"LOGIN ATTEMPT: '{username}'")  

    # ⚠️ التعديل الأهم: زدنا `request` هنا باش django-axes يخدم بشكل صحيح
    user = authenticate(request, username=username, password=password)

    if user is not None:
        print("User found in DB") 
        try:
            profile = user.profile
            print(f"Profile found: Parent={profile.is_parent_account}") 
        except UserProfile.DoesNotExist: # تأكد أن UserProfile مستوردة الفوق
            print("ERROR: Profile NOT FOUND") 
            return Response({"error": "Admin Error: User Profile Missing"}, status=400)

        token, _ = Token.objects.get_or_create(user=user)

        return Response({
            'token': token.key,
            'is_parent': profile.is_parent_account,
            'school_id': profile.school.id,
            'school_name': profile.school.name,
            'school_lat': profile.school.latitude,
            'school_lng': profile.school.longitude,
            'radius': profile.school.radius,
            'admin_badge': profile.school.admin_badge_code,
            'security_pin': profile.school.security_pin,
            'school_api_key': profile.school.api_key,
        })
    
    else:
        # إلا بقى كيعطي خطأ، هاد الـ print غتبين ليك واش المشكل في الباسورد أو اليوزرنيم
        print(f"FAILED AUTH FOR: '{username}'") 
        return Response({"error": "Identifiants incorrects"}, status=400)
        

    
 
# دالة مساعدة باش نجيبو مدرسة المستخدم الحالي
def get_current_school(user):
    try:
        return user.profile.school
    except UserProfile.DoesNotExist:
        return None

# ==========================================
# 1. WEB MONITOR (PROTECTED) 
# ==========================================

@api_view(['GET'])
@permission_classes([AllowAny]) #   ضروري يكون مكونيكطي
def monitor_page(request, school_key):
    # كنقلبو على المدرسة بهاد الكود
    school = get_object_or_404(School, api_key=school_key)

    # كنجيبو أقسام هاد المدرسة فقط
    classes = Student.objects.filter(school=school)\
                             .values_list('class_name', flat=True)\
                             .distinct()\
                             .order_by('class_name')
                             
    garde_categories = GardeCategory.objects.filter(school=school)

    # 2. Nsawbo "Map" bach nwariw l JavaScript chkon classes li f kol Garde
    # Resultat ghadi ykon b7al hakka: { '1': ['PS-A', 'PS-B'], '2': ['CP-A'] }
    garde_map = {}
    for g in garde_categories:
        # Njibo smiyat classes li tab3in l had garde
        classes_names = list(g.classes.values_list('name', flat=True)) 
        garde_map[g.id] = classes_names

    return render(request, 'monitor.html', {
        'classes': classes,
        'school': school,      # باش نكتبو سميتها
        'school_key': school_key,
        'garde_categories': garde_categories, # Bach nrasmo les boutons
        'garde_map_json': json.dumps(garde_map),# باش JS يستعملو من بعد
    }) 

def get_monitor_data(request, school_key):
    school = get_object_or_404(School, api_key=school_key)

    now = timezone.now()
    start_of_day = now.replace(hour=0, minute=0, second=0, microsecond=0)

    selected_classes_str = request.GET.get('classes', '')
    time_slot = request.GET.get('time_slot', 'all')

    query = (
        PickupRequest.objects
        .filter(created_at__gte=start_of_day, student__school=school)
        .select_related('student')
        .order_by('created_at')
    )

    if time_slot == 'morning':
        query = query.filter(created_at__hour__lt=14)
    elif time_slot == 'afternoon':
        query = query.filter(created_at__hour__gte=14)

    if selected_classes_str:
        selected_classes = selected_classes_str.split(',')
        query = query.filter(student__class_name__in=selected_classes)

    unique_requests = []
    seen_student_ids = set()

    for req in query:
        if req.student.id in seen_student_ids:
            continue

        # ✅ porte key
        porte_key = getattr(req, 'porte', None)

        # ✅ 1) porte_label if stored in PickupRequest
        porte_label = getattr(req, 'porte_label', None)

        # ✅ 2) fallback: fetch from SchoolPort via key
        if (not porte_label) and porte_key:
            p = SchoolPort.objects.filter(school=school, key=porte_key).first()
            porte_label = p.label if p else porte_key 

        unique_requests.append({
            'time': req.created_at.strftime("%H:%M:%S"),
            'student_name': f"{req.student.first_name} {req.student.last_name}",
            'class_name': req.student.class_name,
            'is_completed': req.is_completed,

            # ✅ NEW
            'porte': porte_key,
            'porte_label': porte_label,
            'device_id': getattr(req, 'device_id', None),
        })

        seen_student_ids.add(req.student.id)

    unique_requests.reverse()

    # ✅ RETOUR AVEC LE COMPTEUR (COUNT)
    return JsonResponse({
        'requests': unique_requests,
        'count': len(unique_requests)  # ✅ Hahowa l'nombre total
    })

def get_monitor_cdnpdata(request, school_key):
    school = get_object_or_404(School, api_key=school_key)

    now = timezone.now()
    start_of_day = now.replace(hour=0, minute=0, second=0, microsecond=0)

    selected_classes_str = request.GET.get('classes', '')
    time_slot = request.GET.get('time_slot', 'all')

    query = PickupRequest.objects.filter(created_at__gte=start_of_day, student__school=school) \
        .select_related('student') \
        .order_by('created_at')

    if time_slot == 'morning':
        query = query.filter(created_at__hour__lt=14)
    elif time_slot == 'afternoon':
        query = query.filter(created_at__hour__gte=14)

    if selected_classes_str:
        selected_classes = selected_classes_str.split(',')
        query = query.filter(student__class_name__in=selected_classes)

    unique_requests = []
    seen_student_ids = set()

    for req in query:
        if req.student.id not in seen_student_ids:
            unique_requests.append({
                'num': req.student.code_id if req.student.code_id else None,
                'classe': req.student.class_name if req.student.class_name else None,
                'groupe': req.student.class_name if req.student.class_name else None,
                'nom': req.student.first_name if req.student.first_name else None,
                'prenom': req.student.last_name if req.student.last_name else None,
                'badge': req.student.cndp if req.student.cndp else None,
                'badge_str': req.student.badge_number if req.student.badge_number else None,
                'timestamp': None, 'datetime': None, 'date': None,

                # ✅ NEW
                'porte': getattr(req, 'porte', None),
                'device_id': getattr(req, 'device_id', None),
            })
            seen_student_ids.add(req.student.id)

    unique_requests.reverse()
    return JsonResponse({'data': unique_requests})

# ==========================================
# 2. SYNC MOBILE (PROTECTED)
# ==========================================

@api_view(['GET'])
@permission_classes([IsAuthenticated]) #  
def check_db_version(request):
    current_school = get_current_school(request.user)
    if not current_school: return Response({"error": "No School"}, status=403)

    #  غير تلاميذ مدرستو
    students_data = list(Student.objects.filter(school=current_school).values('code_id', 'first_name', 'last_name', 'class_name').order_by('code_id'))
    
    data_string = json.dumps(students_data, cls=DjangoJSONEncoder, sort_keys=True)
    db_hash = hashlib.md5(data_string.encode('utf-8')).hexdigest()
    return Response({'db_hash': db_hash})


@api_view(['GET'])
@permission_classes([IsAuthenticated]) #  
def get_all_students(request):
    current_school = get_current_school(request.user)
    if not current_school: return Response({"error": "No School"}, status=403)

    #  غير تلاميذ مدرستو
    students = Student.objects.filter(school=current_school)
    serializer = StudentSerializer(students, many=True)
    return Response(serializer.data)


# ==========================================
# 3. SCAN & DEMANDES
# ==========================================


def save_to_json_archive(pickup_request):
    """Sauvegarde les données dans un fichier JSON par école et par jour"""
    try:
        # Nettoyer le nom de l'école pour le dossier
        school_name = str(pickup_request.student.school.name).replace(" ", "_")
        folder_path = os.path.join(settings.BASE_DIR, 'archives', school_name)
        
        # Créer le dossier s'il n'existe pas
        if not os.path.exists(folder_path):
            os.makedirs(folder_path)
        
        # Fichier nommé par la date du jour
        file_name = f"{timezone.now().date()}.json"
        file_path = os.path.join(folder_path, file_name)
        
        # Préparer la donnée
        new_entry = {
            "id": pickup_request.id,
            "student": f"{pickup_request.student.first_name} {pickup_request.student.last_name}",
            "class": pickup_request.student.class_name,
            "badge": pickup_request.student.badge_number,
            "porte": pickup_request.porte_label,
            "device_id": pickup_request.device_id,
            "timestamp": pickup_request.created_at.strftime("%Y-%m-%d %H:%M:%S")
        }

        # Lire l'existant et ajouter
        data = []
        if os.path.exists(file_path):
            with open(file_path, 'r', encoding='utf-8') as f:
                try:
                    data = json.load(f)
                except:
                    data = []

        data.append(new_entry)

        # Sauvegarder
        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=4)
            
    except Exception as e:
        print(f"Erreur Archivage JSON: {str(e)}")
        return False

# 2. Ta fonction de Scan modifiée
@api_view(['POST'])
def record_school_scan(request):
    try:
        student_id = request.data.get('student_id')
        porte = request.data.get('porte') or request.headers.get('X-Porte')
        device_id = request.data.get('device_id') or request.headers.get('X-Device-Id')

        school = request.user.profile.school

        # Ignorer le badge Admin
        if str(student_id) == str(school.admin_badge_code):
            return Response({"message": "Scan Admin ignoré"}, status=200)

        # Rechercher l'élève
        student = Student.objects.filter(code_id=student_id, school=school).first()
        if not student:
            return Response({"error": "Élève introuvable"}, status=404)

        # Récupérer le label de la porte
        porte_label = None
        if porte:
            p = SchoolPort.objects.filter(school=school, key=porte).first()
            porte_label = p.label if p else porte

        # ✅ Création dans la Base de Données
        new_pickup = PickupRequest.objects.create(
            student=student,
            is_completed=True,
            porte=porte,
            porte_label=porte_label,
            device_id=device_id
        )

        # ✅ ARCHIVAGE DANS LE FICHIER JSON
        # Même si quelqu'un fait un .delete() sur la DB, ce fichier reste.
        save_to_json_archive(new_pickup)

        return Response(
            {
                "message": "Scan enregistré et archivé", 
                "student": f"{student.first_name} {student.last_name}",
                "porte": porte_label
            }, 
            status=status.HTTP_201_CREATED
        )

    except Exception as e:
        return Response({"error": str(e)}, status=500)

# ... Imports (IsAuthenticated ضروري)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def create_pickup_request(request):
    try:
        incoming_id = request.data.get('student_id')
        if not incoming_id:
            return Response({"error": "student_id manquant"}, status=400)

        porte = request.data.get('porte') or request.headers.get('X-Porte')
        device_id = request.data.get('device_id') or request.headers.get('X-Device-Id')

        # ✅ قصّ الأطوال باش ما يطيحش DB
        if porte is not None:
            porte = str(porte)[:20]
        if device_id is not None:
            device_id = str(device_id)[:80]

        clean_id = str(incoming_id).strip()

        student = Student.objects.filter(
            Q(code_id=clean_id) |
            Q(badge_number=clean_id) |
            Q(id=int(clean_id) if clean_id.isdigit() else -1)
        ).first()

        if not student:
            return Response({"error": "Élève introuvable"}, status=404)

        # طلب واحد Pending لكل student
        req = PickupRequest.objects.filter(student=student, is_completed=False).first()
        if not req:
            req = PickupRequest.objects.create(
                student=student,
                is_completed=False,
                porte=porte,
                device_id=device_id
            )
            created = True
        else:
            created = False
            changed = False
            if porte and req.porte != porte:
                req.porte = porte
                changed = True
            if device_id and req.device_id != device_id:
                req.device_id = device_id
                changed = True
            if changed:
                req.save(update_fields=['porte', 'device_id'])

        return Response({
            "ok": True,
            "message": "Demande envoyée" if created else "Demande déjà existante",
            "pickup_id": req.id
        }, status=201 if created else 200)

    except (DataError, IntegrityError) as e:
        return Response({"error": "DB error", "detail": str(e)}, status=400)

    except Exception as e:
        traceback.print_exc()
        return Response({"error": "Server error", "detail": str(e)}, status=500)



@api_view(['GET'])
@permission_classes([AllowAny])  # (الحارس كيسول) - نفس ديالك
def get_pending_requests(request):
    current_school = get_current_school(request.user)
    if not current_school:
        return Response([])

    requests = PickupRequest.objects.filter(
        is_completed=False,
        student__school=current_school
    ).order_by('created_at')

    data = []
    for req in requests:
        data.append({
            "id": req.id,
            "student": req.student.code_id,
            'student_name': f"{req.student.last_name} {req.student.first_name}",
            "class_name": req.student.class_name,
            "created_at": req.created_at,

            # ✅ NEW
            "porte": getattr(req, 'porte', None),
            "device_id": getattr(req, 'device_id', None),
        })
    return Response(data)

@api_view(['POST'])
@permission_classes([IsAuthenticated]) # 
def complete_request(request, pk):
    current_school = get_current_school(request.user) 
    try:
        #  لازم الطلب يكون تابع لنفس مدرسة الحارس
        req = PickupRequest.objects.get(pk=pk, student__school=current_school)
        req.is_completed = True
        req.save()
        return Response({"status": "success"})
    except PickupRequest.DoesNotExist:
        return Response({"status": "error"}, status=404)


@api_view(['POST'])
@permission_classes([IsAuthenticated]) #   ضروري يكون الحارس
def clear_daily_scans(request):
    # 1. نجيبو مدرسة الحارس
    try:
        current_school = request.user.profile.school
    except:
        return Response({"error": "Aucune école assignée"}, status=403)

    # 2. نحددو بداية اليوم (باش نمسحو غير ديال اليوم)
    now = timezone.now()
    start_of_day = now.replace(hour=0, minute=0, second=0, microsecond=0)

    # 3. نمسحو الطلبات ديال هاد المدرسة + ديال اليوم
    deleted_count, _ = PickupRequest.objects.filter(
        student__school=current_school,
        created_at__gte=start_of_day
    ).delete()

    return Response({"message": "Historique du jour effacé", "count": deleted_count})

# 1. لائحة التلاميذ (مع البحث)


# --- 1. DASHBOARD ECOLE (Gestion Complète) ---

@login_required
def manage_students(request):
    try:
        
        
        school = request.user.profile.school
   
    
        show_config_modal = False
    
        # ✅ AUTO CREATE PORTS
        if not SchoolPort.objects.filter(school=school).exists():
            SchoolPort.objects.bulk_create([
                SchoolPort(school=school, key="port1", label="Porte 1", order=1, is_active=True),
                SchoolPort(school=school, key="port2", label="Porte 2", order=2, is_active=True),
                SchoolPort(school=school, key="port3", label="Porte 3", order=3, is_active=True),
                SchoolPort(school=school, key="port4", label="Porte 4", order=4, is_active=True),
            ])
    
    
        ports = SchoolPort.objects.filter(
            school=school,
            is_active=True
        ).order_by("order")

    
        # ==========================================
        # === POST ACTIONS ===
        # ==========================================
        if request.method == 'POST':
            action = request.POST.get('action')
    
            # ✅ NEW: RENAME PORTS FROM HTML FORM
            if action == 'rename_port':
                key = request.POST.get('port_key', '').strip()
                label = request.POST.get('port_label', '').strip()
            
                if not key or not label:
                    messages.error(request, "❌ Port key / label manquant.")
                else:
                    updated = SchoolPort.objects.filter(
                        school=school,
                        key=key
                    ).update(label=label[:50])
            
                    if updated:
                        messages.success(request, f"✅ Porte '{key}' renommée en '{label}'.")
                    else:
                        messages.error(request, f"❌ Porte '{key}' introuvable.")
            
                return redirect(request.path)   # ✅ يرجّع نفس الصفحة بلا ما تكرر POST
            
            elif action == 'change_login_password':
                    # print باش تشوف ف الـ terminal واش الدالة خدمات
                    print("--- Action: Change Password Started ---") 
                    
                    new_pass = request.POST.get('new_password')
                    confirm_pass = request.POST.get('confirm_password')
        
                    if new_pass and confirm_pass:
                        if new_pass == confirm_pass:
                            if len(new_pass) < 6:
                                messages.warning(request, "⚠️ Le mot de passe doit contenir au moins 6 caractères.")
                            else:
                                u = request.user
                                u.set_password(new_pass)
                                u.save()
                                # هادي مهمة باش الـ session تبقى خدامة ومايخرجش
                                update_session_auth_hash(request, u) 
                                messages.success(request, "🔐 Mot de passe modifié avec succès !")
                        else:
                            messages.error(request, "❌ Les mots de passe ne correspondent pas.")
                    else:
                        messages.error(request, "❌ Veuillez remplir les champs.")
                    
                    # مهم جداً: باش ملي ترجع الصفحة، تمشي نيشان لـ Tab ديال Localisation
                    # خاصك دوز هاد المتغير context للـ template فاللخر
                    active_tab = 'location' 
        
                # ... (باقي الكود) ...
                
                # فاللخر ملي دير render، تأكد بلي active_tab كاين
                # إلا ماكانش عندك متغير سميتو active_tab قبل، عرفو الفوق بـ 'students'
                # context = { ..., 'active_tab': active_tab if 'active_tab' in locals() else 'students' }
            
    
            # A. UPDATE LOCATION & PIN
            if action == 'update_location':
                new_radius = request.POST.get('radius')
                new_lat = request.POST.get('latitude')
                new_lng = request.POST.get('longitude')
                new_pin = request.POST.get('security_pin')
                is_active = request.POST.get('active_admin_scan')
    
                if new_radius:
                    school.radius = int(new_radius)
    
                if new_lat and new_lng:
                    try:
                        school.latitude = float(new_lat.replace(',', '.'))
                        school.longitude = float(new_lng.replace(',', '.'))
                    except:
                        pass
    
                if new_pin:
                    clean_pin = str(new_pin).strip()
                    if len(clean_pin) == 4 and clean_pin.isdigit():
                        school.security_pin = clean_pin
    
                try:
                    school.active_admin_scan = is_active
                    school.save()
                except:
                    school.active_admin_scan = False
                    school.save()
    
                messages.success(request, "📍 Configuration mise à jour !")
    
            # B. IMPORT CSV
            elif action == 'upload_csv':
                try:
                    csv_file = request.FILES.get('file')
                    if not csv_file.name.endswith('.csv'):
                        messages.error(request, "Le fichier doit être un CSV (.csv)")
                    else:
                        file_data = csv_file.read()
                        try:
                            decoded_file = file_data.decode('utf-8-sig')
                        except UnicodeDecodeError:
                            decoded_file = file_data.decode('latin-1')
    
                        io_string = io.StringIO(decoded_file)
                        first_line = io_string.readline()
                        separator = ';' if ';' in first_line else ','
                        io_string.seek(0)
                        next(io_string, None)
    
                        csv_reader = csv.reader(io_string, delimiter=separator)
                        count = 0
                        for row in csv_reader:
                            try:
                                if len(row) < 7:
                                    continue
                                Student.objects.create(
                                    code_id=row[0].strip(),
                                    class_name=row[1].strip(),
                                    group=row[2].strip(),
                                    last_name=row[3].strip(),
                                    first_name=row[4].strip(),
                                    cndp=row[5].strip(),
                                    badge_number=row[6].strip(),
                                    school=school,
                                )
                                count += 1
                            except:
                                continue
                        messages.success(request, f"✅ {count} élèves ajoutés !")
                except Exception as e:
                    messages.error(request, f"Erreur : {str(e)}")
    
            # B2. IMPORT XLSX
            elif action == 'upload_xlsx':
                try:
                    excel_file = request.FILES.get('file')
    
                    if not excel_file.name.endswith('.xlsx'):
                        messages.error(request, "Le fichier doit être un Excel (.xlsx)")
                    else:
                        try:
                            wb = openpyxl.load_workbook(excel_file)
                            ws = wb.active
    
                            count = 0
                            for row in ws.iter_rows(min_row=2, values_only=True):
                                try:
                                    if not row or len(row) < 7:
                                        continue
    
                                    Student.objects.create(
                                        code_id=str(row[0]).strip() if row[0] else '',
                                        class_name=str(row[1]).strip() if row[1] else '',
                                        group=str(row[2]).strip() if row[2] else '',
                                        last_name=str(row[3]).strip() if row[3] else '',
                                        first_name=str(row[4]).strip() if row[4] else '',
                                        cndp=str(row[5]).strip() if row[5] else '',
                                        badge_number=str(row[6]).strip() if row[6] else '',
                                        school=school,
                                    )
                                    count += 1
                                except:
                                    continue
    
                            messages.success(request, f"✅ {count} élèves ajoutés via Excel !")
    
                        except Exception as e:
                            messages.error(request, f"Erreur de lecture Excel : {str(e)}")
    
                except Exception as e:
                    messages.error(request, f"Erreur globale : {str(e)}")
    
            # C. EXPORT CSV
            elif action == 'export_csv':
                response = HttpResponse(content_type='text/csv')
                response['Content-Disposition'] = 'attachment; filename="eleves_export.csv"'
                response.write(u'\ufeff'.encode('utf8'))
                writer = csv.writer(response, delimiter=';')
                writer.writerow(['code_id','Nom', 'Prénom', 'Classe','group' ,'Numéro Badge', 'Code Massar','school'])
    
                students_export = Student.objects.filter(school=school).order_by('class_name', 'last_name')
                for s in students_export:
                    writer.writerow([s.code_id, s.last_name, s.first_name, s.class_name, s.group, s.badge_number, s.cndp, s.school])
                return response
            
            elif action == 'export_xlsx':
            
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Liste des Elèves"
            
                # 1. تعريف الألوان والتنسيقات
                header_fill = PatternFill(start_color="FFC107", end_color="FFC107", fill_type="solid") # لون أصفر وورنينج
                header_font = Font(bold=True, color="FFFFFF", size=12)
                center_alignment = Alignment(horizontal="center", vertical="center")
                border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            
                # 2. تحديد الرؤوس
                headers = ['Nom', 'Prénom', 'Classe', 'Group', 'Numéro Badge', 'Code Massar']
                ws.append(headers)
            
                # تنسيق سطر الرؤوس
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center_alignment
                    cell.border = border
            
                # 3. إضافة البيانات
                students_export = Student.objects.filter(school=school).order_by('class_name', 'last_name')
                for s in students_export:
                    ws.append([s.last_name, s.first_name, s.class_name, s.group, s.badge_number, s.cndp])
            
                # 4. تفعيل الحماية وتنسيق الأعمدة
                ws.protection.sheet = True
                ws.protection.password = '123'
                
                unlocked = Protection(locked=False)
                locked = Protection(locked=True)
                badge_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid") # لون رمادي للتمييز
            
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    for cell in row:
                        cell.border = border
                        cell.alignment = Alignment(horizontal="left")
                        
                        if cell.column == 5: # عمود Numéro Badge
                            cell.protection = locked
                            cell.fill = badge_fill # تلوينه بالرمادي ليعرف المستخدم أنه للقراءة فقط
                        else:
                            cell.protection = unlocked
            
                # ضبط عرض الأعمدة تلقائياً ليكون الشكل مرتباً
                column_widths = [20, 20, 15, 10, 15, 20]
                for i, width in enumerate(column_widths, 1):
                    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
            
                response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = f'attachment; filename="Eleves_{school.name}.xlsx"'
                wb.save(response)
                return response
            
            
            elif action == 'upload_xlsxV2':
                try:
                    excel_file = request.FILES.get('file')
                    if not excel_file or not excel_file.name.endswith('.xlsx'):
                        messages.error(request, "Veuillez sélectionner un fichier Excel (.xlsx)")
                    else:
                        wb = openpyxl.load_workbook(excel_file)
                        ws = wb.active
                        
                        count_updated = 0
                        not_found_badges = [] # قائمة لتخزين البادجات غير الموجودة
            
                        # نمر على الصفوف (البيانات تبدأ من الصف 2)
                        for row in ws.iter_rows(min_row=2, values_only=True):
                            # الترتيب: 0:Nom, 1:Prénom, 2:Classe, 3:Group, 4:Badge, 5:Massar
                            b_number = str(row[4]).strip() if row[4] else None
                            
                            if not b_number:
                                continue
            
                            # البحث عن التلميذ
                            student = Student.objects.filter(school=school, badge_number=b_number).first()
            
                            if student:
                                # تحديث البيانات
                                student.last_name = str(row[0]).strip() if row[0] else student.last_name
                                student.first_name = str(row[1]).strip() if row[1] else student.first_name 
                                student.class_name = str(row[2]).strip() if row[2] else student.class_name
                                student.group = str(row[3]).strip() if row[3] else student.group
                                student.cndp = str(row[5]).strip() if row[5] else student.cndp
                                student.save()
                                count_updated += 1
                            else:
                                # إذا لم يوجد، نضيف رقم البادج للقائمة لإعلام المستخدم
                                not_found_badges.append(b_number)
            
                        # رسالة النجاح
                        #messages.success(request, f"✅ تم تحديث {count_updated} تلميذ بنجاح.")
                        
                        
                        messages.success(request, f"✅{count_updated} élèves ont été mis à jour avec succès.")
                        # رسالة التنبيه في حال وجود بادجات مفقودة
                        if not_found_badges:
                            badges_str = ", ".join(not_found_badges)
                            #messages.warning(request, f"⚠️ البادجات التالية غير موجودة في قاعدة البيانات ولم يتم تحديثها: {badges_str}")
                            messages.warning(request, f"⚠️ Les badges suivants ne figurent pas dans la base de données et n'ont pas été mis à jour:{badges_str}")
                except Exception as e:
                    messages.error(request, f"Erreur lors de l'import : {str(e)}")
            
            # D. CLEAR LIST
            elif action == 'clear_list':
                Student.objects.filter(school=school).delete()
                messages.warning(request, "🗑️ Liste vidée.")
                
            
            elif action == 'create_reserved':
                try:
                    # --- A. حساب NUMÉRO DE BADGE (Max + 1) ---
                    # 1. نجلب كل أرقام البادجات لهاته المدرسة
                    existing_badges = Student.objects.filter(school=school).values_list('badge_number', flat=True)
                    
                    max_badge = 0
                    for b in existing_badges:
                        # نتأكد أن البادج رقم فقط (نتجاهل RESERVE أو A12)
                        if b and str(b).isdigit():
                            val = int(b)
                            if val > max_badge:
                                max_badge = val
                    
                    # البادج الجديد هو الأكبر + 1
                    next_badge_number = str(max_badge + 1)
    
    
                    # --- B. حساب CODE ID (YYYY YY ID + 001) ---
                    
                    yyyy = datetime.now().year
                    #yyyy = now.strftime("%y")       # "25"
                    school_key = str(school.id)   # "1"
    
                    # الصيغة: 25 + 25 + 1 = 25251
                    prefix_str = f"{yyyy}{yyyy}{school_key}" 
                    
                    # نضرب في 1000 لنترك مساحة للعداد
                    # 25251000
                    base_calculation = int(prefix_str) * 1000
    
                    counter = 1
                    while True:
                        # 25251000 + 1 = 25251001
                        final_result = base_calculation + counter
                        candidate_code = str(final_result)
                        
                        if not Student.objects.filter(code_id=candidate_code).exists():
                            break 
                        counter += 1
    
                    # 4. إنشاء التلميذ
                    Student.objects.create(
                        school=school,
                        first_name="RESERVE",
                        last_name="RESERVE",
                        class_name="RESERVE",
                        group ="RESERVE",
                        # ✅ هنا كنحطو الرقم التسلسلي الجديد (مثلاً 150)
                        badge_number=next_badge_number, 
                        
                        cndp="0000",
                        code_id=candidate_code 
                    )
    
                    messages.success(request, f"✅ Réservé : Badge N° {next_badge_number} (ID: {candidate_code})")
    
                except Exception as e:
                    messages.error(request, f"Erreur Création : {str(e)}")
           
            elif action == 'create_Mreserved':
                try:
                    # 1. جلب العدد المطلوب
                    try:
                        count_to_create = int(request.POST.get('count', 1))
                    except ValueError:
                        count_to_create = 1
            
                    # السنة بـ 2 أرقام (مثلاً 26)
                    yy_short = datetime.now().strftime('%y') 
                    school_key = str(school.api_key)
                    
                    # الـ Prefix دابا غيكون فيه غير 26 + مفتاح المدرسة
                    badge_prefix = f"{yy_short}{school_key}"
                    
                    # --- تحسين الأداء (Optimization) ---
                    existing_badges = set(Student.objects.filter(
                        school=school, 
                        badge_number__startswith=badge_prefix
                    ).values_list('badge_number', flat=True))
                    
                    existing_codes = set(Student.objects.filter(
                        school=school
                    ).values_list('code_id', flat=True))
            
                    current_seq = 1
                    created_count = 0
            
                    # --- حلقة التكرار ---
                    for _ in range(count_to_create):
                        
                        # 1. البحث عن Badge Number متاح
                        while True:
                            # zfill(2) هي اللي كدير لينا 01, 02 عوض 1, 2
                            seq_formatted = str(current_seq).zfill(2)
                            candidate_badge = f"{badge_prefix}{seq_formatted}"
                            
                            if candidate_badge not in existing_badges:
                                break
                            current_seq += 1
                        
                        # 2. لوجيك الـ Code ID (ديالك)
                        # كنستعملو السنة كاملة 2026 في الحساب باش يبقى الرقم فريد
                        yyyy_full = datetime.now().year
                        codage = yyyy_full * current_seq
                        prefix_str = f"{yyyy_full}{school_key}{codage}"
                        base_calculation = int(prefix_str) * 1000
                        
                        counter = 1
                        while True:
                            candidate_code = str(base_calculation + counter)
                            if candidate_code not in existing_codes:
                                break 
                            counter += 1
            
                        # 3. إنشاء السجل
                        Student.objects.create(
                            school=school,
                            first_name="RESERVE",
                            last_name="RESERVE",
                            class_name="RESERVE",
                            group="RESERVE",
                            badge_number=candidate_badge,
                            cndp="0000",
                            code_id=candidate_code 
                        )
                        
                        # تحديث الـ Sets
                        existing_badges.add(candidate_badge)
                        existing_codes.add(candidate_code)
                        
                        current_seq += 1
                        created_count += 1
            
                    messages.success(request, f"{created_count} Badge(s) de reserve créé avec succès.")
                except Exception as e:
                    messages.error(request, f"Erreur Création : {str(e)}")
            
            
            elif action == 'print_badges':
                response = HttpResponse(content_type='application/pdf')
                response['Content-Disposition'] = 'attachment; filename="badges_securises.pdf"'
    
                p = canvas.Canvas(response, pagesize=A4)
                width, height = A4
                
                margin_x = 1 * cm
                margin_y = 1 * cm
                col_width = (width - 2 * margin_x) / 3
                row_height = (height - 2 * margin_y) / 4
    
                students = Student.objects.filter(school=school).order_by('class_name', 'last_name')
    
                col = 0
                row = 3 
                
                # ✅ لائحة الثوابت (نفسها لي فـ Flutter)
                CONSTANTS = [1313131313, 1234567899, 1234567899, 1234567899, 1234567899, 
                             1234567899, 1234567899]
    
                for s in students:
                    x = margin_x + col * col_width
                    y = margin_y + row * row_height
    
                    # Cadre
                    p.setStrokeColorRGB(0.8, 0.8, 0.8)
                    p.rect(x + 0.2*cm, y + 0.2*cm, col_width - 0.4*cm, row_height - 0.4*cm)
    
                    # 1. Numéro Badge (Clair)
                    p.setFont("Helvetica-Bold", 16)
                    display_badge = s.badge_number if s.badge_number else "---"
                    p.drawCentredString(x + col_width/2, y + row_height - 1.5*cm, f"N° {display_badge}")
    
                    # 2. ✅✅✅ CALCUL DU CRYPTAGE (ENCODAGE)
                    try:
                        # نحولو ID لرقم
                        val = int(s.code_id)
                        
                        # الضرب فـ 13 (10 مرات)
                        val = val * (13**10)
                        
                        # زيادة الثوابت 
                        for num in CONSTANTS:
                            val = val + num
                        
                        encrypted_code = str(val)
                    except:
                        # إلا كان الكود فيه حروف، كنخليوه كما هو
                        encrypted_code = encrypted_code
    
                    # 3. QR Code (فيه الكود المشفر)
                    qr = qrcode.QRCode(box_size=10, border=1)
                    qr.add_data(encrypted_code) # 👈 هنا كنحطو الكود الطويل
                    qr.make(fit=True)
                    img_qr = qr.make_image(fill_color="black", back_color="white")
    
                    img_buffer = io.BytesIO()
                    img_qr.save(img_buffer, format='PNG')
                    img_buffer.seek(0)
                    
                    qr_size = 4.5 * cm 
                    qr_y = y + (row_height - qr_size) / 2 - 0.5 * cm 
                    p.drawImage(ImageReader(img_buffer), x + col_width/2 - qr_size/2, qr_y, width=qr_size, height=qr_size)
    
                    col += 1
                    if col >= 3:
                        col = 0
                        row -= 1
                        if row < 0: 
                            p.showPage()
                            row = 3
                
                p.save()
                return response
            
            # ... (dakchi li 9bel)
    
            elif action == 'print_badges_filter':
                # 1. Recuperation des filtres depuis le POST (Hidden inputs)
                filter_class = request.POST.get('filter_class', '').strip()
                filter_q = request.POST.get('filter_q', '').strip()
    
                # 2. Base Query
                students = Student.objects.filter(school=school)
    
                # 3. Appliquer le filtre Classe si existe
                if filter_class:
                    students = students.filter(class_name=filter_class)
    
                # 4. Appliquer le filtre Recherche (kif ma derti f GET)
                if filter_q:
                    students = students.filter(
                        Q(last_name__icontains=filter_q) | 
                        Q(first_name__icontains=filter_q) | 
                        Q(code_id__icontains=filter_q) |
                        Q(badge_number__icontains=filter_q)
                    )
    
                # 5. Ordre
                students = students.order_by('class_name', 'last_name')
    
                # --- GENERATION PDF (Code dyalk l9dim b9a howa howa) ---
                response = HttpResponse(content_type='application/pdf')
                response['Content-Disposition'] = 'attachment; filename="badges_securises.pdf"'
    
                p = canvas.Canvas(response, pagesize=A4)
                width, height = A4
                
                margin_x = 1 * cm
                margin_y = 1 * cm
                col_width = (width - 2 * margin_x) / 3
                row_height = (height - 2 * margin_y) / 4
    
                col = 0
                row = 3 
                
                CONSTANTS = [1313131313, 1234567899, 1234567899, 1234567899, 1234567899, 
                             1234567899, 1234567899]
    
                # Vérification : Ila makayninach talamid
                if not students.exists():
                    messages.warning(request, "Aucun élève à imprimer avec ces filtres.")
                    return redirect('manage_students')
    
                for s in students:
                    # ... (Nfs lcode dyal rasm lbadge li kan 3ndk) ...
                    # ... Met lcode l9dim hna ...
                    x = margin_x + col * col_width
                    y = margin_y + row * row_height
    
                    p.setStrokeColorRGB(0.8, 0.8, 0.8)
                    p.rect(x + 0.2*cm, y + 0.2*cm, col_width - 0.4*cm, row_height - 0.4*cm)
    
                    p.setFont("Helvetica-Bold", 16)
                    display_badge = s.badge_number if s.badge_number else "---"
                    p.drawCentredString(x + col_width/2, y + row_height - 1.5*cm, f"N° {display_badge}")
    
                    try:
                        val = int(s.code_id)
                        val = val * (13**10)
                        for num in CONSTANTS:
                            val = val + num
                        encrypted_code = str(val)
                    except:
                        encrypted_code = s.code_id # Fallback
    
                    qr = qrcode.QRCode(box_size=10, border=1)
                    qr.add_data(encrypted_code)
                    qr.make(fit=True)
                    img_qr = qr.make_image(fill_color="black", back_color="white")
    
                    img_buffer = io.BytesIO()
                    img_qr.save(img_buffer, format='PNG')
                    img_buffer.seek(0)
                    
                    qr_size = 4.5 * cm 
                    qr_y = y + (row_height - qr_size) / 2 - 0.5 * cm 
                    p.drawImage(ImageReader(img_buffer), x + col_width/2 - qr_size/2, qr_y, width=qr_size, height=qr_size)
    
                    col += 1
                    if col >= 3:
                        col = 0
                        row -= 1
                        if row < 0: 
                            p.showPage()
                            row = 3 
                
                p.save()
                return response
            
            
            elif action == 'download_zip_badgesbkp':
                # 1. Recuperation des filtres (nfs l-manti9)
                filter_class = request.POST.get('filter_class', '').strip()
                filter_q = request.POST.get('filter_q', '').strip()
            
                students = Student.objects.filter(school=school)
                if filter_class:
                    students = students.filter(class_name=filter_class)
                if filter_q: 
                    students = students.filter(
                        Q(last_name__icontains=filter_q) | 
                        Q(first_name__icontains=filter_q) | 
                        Q(code_id__icontains=filter_q) |
                        Q(badge_number__icontains=filter_q)
                    )
            
                if not students.exists():
                    messages.warning(request, "Aucun élève trouvé.")
                    return redirect('manage_students')
            
                # 2. Khl9 l-buffer dyal l-ZIP
                zip_buffer = io.BytesIO()
            
                with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                    CONSTANTS = [1313131313, 1234567899, 1234567899, 1234567899, 1234567899, 1234567899, 1234567899]
            
                    for s in students:
                        # Khl9 buffer l kol PDF dyal tilmid
                        pdf_buffer = io.BytesIO()
                        # Ghadi nkhdmo b size sghir (masalan 10cm x 10cm) hit kolla PDF fih QR wahd
                        # Ila bghiti A4 khliha A4
                        p = canvas.Canvas(pdf_buffer, pagesize=(10*cm, 10*cm))
                        
                        # --- CALCUL DU CRYPTAGE ---
                        try:
                            val = int(s.code_id)
                            val = val * (13**10)
                            for num in CONSTANTS:
                                val = val + num
                            encrypted_code = str(val)
                        except:
                            encrypted_code = s.code_id
            
                        # --- GENERATION QR ---
                        qr = qrcode.QRCode(box_size=10, border=1)
                        qr.add_data(encrypted_code)
                        qr.make(fit=True)
                        img_qr = qr.make_image(fill_color="black", back_color="white")
            
                        qr_temp_buffer = io.BytesIO()
                        img_qr.save(qr_temp_buffer, format='PNG')
                        qr_temp_buffer.seek(0)
            
                        # --- RASM F PDF ---
                        p.setFont("Helvetica-Bold", 12)
                        display_name = f"SCAN ME "
                        p.drawCentredString(5*cm, 8.5*cm, display_name)
                        p.drawCentredString(5*cm, 1.5*cm, f"N° {s.badge_number or '---'}")
                        
                        p.drawImage(ImageReader(qr_temp_buffer), 2.5*cm, 2.5*cm, width=5*cm, height=5*cm)
                        
                        p.showPage()
                        p.save()
            
                        # 3. Zid l-PDF l l-ZIP
                        pdf_buffer.seek(0)
                        file_name = f"badge_{s.badge_number or s.id}_{s.last_name}.pdf"
                        zip_file.writestr(file_name, pdf_buffer.read())
            
                # 4. Erja3 l-ZIP kaml l-user
                zip_buffer.seek(0)
                response = HttpResponse(zip_buffer.read(), content_type='application/zip')
                response['Content-Disposition'] = 'attachment; filename="badges_eleves.zip"'
                return response
            
            elif action == 'download_zip_badges':
                # 1. Recuperation des filtres (nfs l-manti9)
                filter_class = request.POST.get('filter_class', '').strip()
                filter_q = request.POST.get('filter_q', '').strip()
            
                students = Student.objects.filter(school=school)
                if filter_class:
                    students = students.filter(class_name=filter_class)
                if filter_q: 
                    students = students.filter(
                        Q(last_name__icontains=filter_q) | 
                        Q(first_name__icontains=filter_q) | 
                        Q(code_id__icontains=filter_q) |
                        Q(badge_number__icontains=filter_q)
                    )
            
                if not students.exists():
                    messages.warning(request, "Aucun élève trouvé.")
                    return redirect('manage_students')
            
                # 2. Khl9 l-buffer dyal l-ZIP
                zip_buffer = io.BytesIO()
            
                with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                    
            
                    for s in students:
                        # Khl9 buffer l kol PDF dyal tilmid
                        pdf_buffer = io.BytesIO()
                        # Ghadi nkhdmo b size sghir (masalan 10cm x 10cm) hit kolla PDF fih QR wahd
                        # Ila bghiti A4 khliha A4
                        p = canvas.Canvas(pdf_buffer, pagesize=(10*cm, 10*cm))
                        
                        # --- CALCUL DU CRYPTAGE ---
                        try:
                            val = int(s.code_id)
                        
                            
                            encrypted_code = str(val)
                        except:
                            encrypted_code = s.code_id
            
                        # --- GENERATION QR ---
                        qr = qrcode.QRCode(box_size=10, border=1)
                        qr.add_data(encrypted_code)
                        qr.make(fit=True)
                        img_qr = qr.make_image(fill_color="black", back_color="white")
            
                        qr_temp_buffer = io.BytesIO()
                        img_qr.save(qr_temp_buffer, format='PNG')
                        qr_temp_buffer.seek(0)
            
                        # --- RASM F PDF ---
                        p.setFont("Helvetica-Bold", 12)
                        display_name = f"SCAN ME "
                        p.drawCentredString(5*cm, 8.5*cm, display_name)
                        p.drawCentredString(5*cm, 1.5*cm, f"N° {s.badge_number or '---'}")
                        
                        p.drawImage(ImageReader(qr_temp_buffer), 2.5*cm, 2.5*cm, width=5*cm, height=5*cm)
                        
                        p.showPage()
                        p.save()
            
                        # 3. Zid l-PDF l l-ZIP
                        pdf_buffer.seek(0)
                        file_name = f"badge_{s.badge_number or s.id}_{s.last_name}.pdf"
                        zip_file.writestr(file_name, pdf_buffer.read())
            
                # 4. Erja3 l-ZIP kaml l-user
                zip_buffer.seek(0)
                response = HttpResponse(zip_buffer.read(), content_type='application/zip')
                response['Content-Disposition'] = 'attachment; filename="badges_eleves.zip"'
                return response
            
            
            elif action == 'delete_student':
                try:
                    student_id = request.POST.get('student_id')
                    
                    # البحث عن التلميذ (مع التأكد أنه تابع لنفس المدرسة للحماية)
                    student = Student.objects.get(id=student_id, school=school)
                    
                    # الحذف
                    name = f"{student.first_name} {student.last_name}"
                    student.delete()
                    
                    messages.success(request, f"🗑️ L'élève {name} a été supprimé.")
                    
                except Student.DoesNotExist:
                    messages.error(request, "❌ Élève introuvable.")
                except Exception as e:
                    messages.error(request, f"Erreur : {str(e)}")
           
            elif action == 'add_garde_category':
                cat_name = request.POST.get('category_name')
                if cat_name:
                    GardeCategory.objects.create(school=school, name=cat_name)
                    messages.success(request, f"✅ Catégorie '{cat_name}' créée !")
    
            # B. AJOUTER/MODIFIER CLASSE AVEC GARDE
            elif action == 'add_class':
                class_name = request.POST.get('new_class_name', '').strip().upper()
                garde_id = request.POST.get('garde_id') # <-- Hada jdid
                
                if class_name:
                    # Njibo Garde object ila kan choisi
                    selected_garde = None
                    if garde_id:
                        try:
                            selected_garde = GardeCategory.objects.get(id=garde_id, school=school)
                        except:
                            pass
    
                    # N9albo wach classe kayna wla la (Update or Create)
                    obj, created = SchoolClass.objects.update_or_create(
                        school=school, 
                        name=class_name,
                        defaults={'garde_category': selected_garde} # <-- Hna kanls9o l garde
                    )
                    
                    if created:
                        messages.success(request, f"✅ Classe {class_name} ajoutée avec succès !")
                    else:
                        messages.info(request, f"🔄 Classe {class_name} mise à jour.")
                    show_config_modal = True
                
               
           
           # ... (dakchi li kan 9bel: add_garde_category, add_class...)
    
            # 3. MODIFIER UNE GARDE
            elif action == 'edit_garde':
                garde_id = request.POST.get('garde_id')
                new_name = request.POST.get('new_name')
                if garde_id and new_name:
                    try:
                        garde = GardeCategory.objects.get(id=garde_id, school=school)
                        garde.name = new_name
                        garde.save()
                        messages.success(request, f"✅ Garde renommée en '{new_name}'.")
                    except GardeCategory.DoesNotExist:
                        messages.error(request, "Erreur: Garde introuvable.")
                show_config_modal = True
    
            # 4. SUPPRIMER UNE GARDE
            elif action == 'delete_garde':
                garde_id = request.POST.get('garde_id')
                if garde_id:
                    try:
                        garde = GardeCategory.objects.get(id=garde_id, school=school)
                        garde_name = garde.name
                        garde.delete()
                        # Melli katmsse7, les classes kaywlliw NULL automatiquement hit drna on_delete=SET_NULL
                        messages.warning(request, f"🗑️ La catégorie '{garde_name}' a été supprimée.")
                    except:
                        messages.error(request, "Erreur lors de la suppression.")
                show_config_modal = True
    
            # 5. ANNULER L'AFFECTATION (Détacher une classe)
            elif action == 'remove_class_garde':
                class_id = request.POST.get('class_id')
                if class_id:
                    try:
                        cls = SchoolClass.objects.get(id=class_id, school=school)
                        cls.garde_category = None # Hna kanraj3oha khawya
                        cls.save()
                        messages.info(request, f"🔗 Garde retirée pour la classe {cls.name}.")
                    except:
                        messages.error(request, "Erreur: Classe introuvable.")
                        
                show_config_modal = True
           
    
            # 6. CRÉER UNE CLASSE SEULEMENT (Sans Garde)
            elif action == 'create_simple_class':
                class_name = request.POST.get('class_name', '').strip().upper()
                if class_name:
                    # get_or_create: Ila kant kayjibha, ila makantch kaycréerha
                    obj, created = SchoolClass.objects.get_or_create(school=school, name=class_name)
                    
                    if created:
                        messages.success(request, f"✅ Classe '{class_name}' créée.")
                    else:
                        messages.warning(request, f"⚠️ La classe '{class_name}' existe déjà.")
                
                show_config_modal = True # Bach tb9a lmodal m7loula
            
            # F west manage_students (m3a actions POST)
    
            elif action == 'assign_class_garde':
                try:
                    class_id = request.POST.get('class_id') # ID dyal classe li bghina nzido
                    garde_id = request.POST.get('garde_id') # ID dyal Garde fin bghina nkhchiwha
                    
                    school_class = get_object_or_404(SchoolClass, id=class_id, school=school)
                    garde = get_object_or_404(GardeCategory, id=garde_id, school=school)
                    
                    school_class.garde_category = garde
                    school_class.save()
                    
                    messages.success(request, f"Classe {school_class.name} ajoutée à {garde.name}")
                except Exception as e:
                    messages.error(request, f"Erreur d'affectation : {str(e)}")
        
            elif action == 'remove_class_garde':
                try:
                    class_id = request.POST.get('class_id')
                    school_class = get_object_or_404(SchoolClass, id=class_id, school=school)
                    
                    old_garde = school_class.garde_category.name if school_class.garde_category else ""
                    school_class.garde_category = None
                    school_class.save()
                    
                    messages.success(request, f"Classe {school_class.name} retirée de {old_garde}")
                except Exception as e:
                    messages.error(request, f"Erreur : {str(e)}")
            
            elif action == 'delete_selected':
                selected_ids = request.POST.getlist('student_ids')
                confirm_pass = request.POST.get('confirm_pass_input') # تأكد أن هذا الاسم يطابق الـ JS
            
                if confirm_pass == "0000":
                    if selected_ids:
                        # تنفيذ الحذف
                        deleted_count = Student.objects.filter(id__in=selected_ids, school=school).delete()[0]
                        messages.success(request, f"✅ {deleted_count} élève(s) supprimé(s) avec succès.")
                    else:
                        messages.warning(request, "⚠️ Veuillez sélectionner au moins un élève.")
                else:
                    # هذه الرسالة التي ظهرت لك تعني أن confirm_pass لم تكن "000"
                    messages.error(request, "❌ Mot de passe de confirmation incorrect. Suppression annulée.")
                    
            if action in ['add_garde_category', 'add_configured_class', 'edit_garde', 'delete_garde', 'remove_class_garde']:
                
                # ... (Lcode li kan 3ndk hna dyal traitemement khelih kif ma howa) ...
                
                # Zid hadi f kol block, awla f lakher d had actions:
                show_config_modal = True
           
            
    
            # ✅ باقي actions ديالك (create_reserved, print_badges, zip, delete_student, garde, class...)
            # خليتهم كيف ما هما، ما بدلت فيهم والو (نفس الكود ديالك)
    
            # ... continue your existing blocks here unchanged ...
    
    
        # ==========================================
        # === GET DISPLAY ===
        # ==========================================
    
        students_list = Student.objects.filter(school=school)\
            .annotate(cleaned_class=Trim(Lower('class_name')))\
            .order_by('cleaned_class', 'last_name')
    
        # a. classes من students
        classes_from_students = set(
            Student.objects.filter(school=school)
            .exclude(class_name__isnull=True)
            .exclude(class_name__exact='')
            .values_list('class_name', flat=True)
        )
    
        # b. classes من config
        classes_from_config = set(
            SchoolClass.objects.filter(school=school)
            .values_list('name', flat=True)
        )
    
        # c. union
        all_classes_set = classes_from_students.union(classes_from_config)
        all_classes = list(all_classes_set)
    
        saved_order_str = school.class_order
        if saved_order_str:
            saved_order = [x.strip() for x in saved_order_str.split(',') if x.strip()]
            order_map = {name: i for i, name in enumerate(saved_order)}
            all_classes.sort(key=lambda x: order_map.get(x, 999))
        else:
            all_classes.sort()
    
        garde_categories = GardeCategory.objects.filter(school=school)
        configured_classes = SchoolClass.objects.filter(school=school).select_related('garde_category').order_by('name')
        student_classes_list = all_classes
    
        search_query = request.GET.get('q', '').strip()
        selected_class = request.GET.get('classe', '').strip()
    
        if search_query:
            students_list = students_list.filter(
                Q(last_name__icontains=search_query) |
                Q(first_name__icontains=search_query) |
                Q(code_id__icontains=search_query) |
                Q(badge_number__icontains=search_query)
            )
    
        if selected_class:
            students_list = students_list.filter(class_name=selected_class)
    
        per_page = request.GET.get('per_page', '100')
        try:
            per_page = int(per_page)
        except:
            per_page = 100
    
        paginator = Paginator(students_list, per_page)
        page_number = request.GET.get('page')
        students = paginator.get_page(page_number)
    
        # ✅ IMPORTANT: ports خاصها تتبعث هنا للـ template
        return render(request, 'manage_students.html', {
            'school': school,
            'students': students,
            'per_page': per_page,
            'all_classes': all_classes,
            'search_query': search_query,
            'selected_class': selected_class,
            'garde_categories': garde_categories,
            'configured_classes': configured_classes,
            'student_classes_list': student_classes_list,
            'show_config_modal': show_config_modal,
    
            # ✅ NEW
            'ports': ports,
        })

    except:
        return render(request, 'manage_students.html')


@login_required
def manage_admin(request):
    try:
        
        
        school = request.user.profile.school
   
    
        show_config_modal = False
    
        # ✅ AUTO CREATE PORTS
        if not SchoolPort.objects.filter(school=school).exists():
            SchoolPort.objects.bulk_create([
                SchoolPort(school=school, key="port1", label="Porte 1", order=1, is_active=True),
                SchoolPort(school=school, key="port2", label="Porte 2", order=2, is_active=True),
                SchoolPort(school=school, key="port3", label="Porte 3", order=3, is_active=True),
                SchoolPort(school=school, key="port4", label="Porte 4", order=4, is_active=True),
            ])
    
    
        ports = SchoolPort.objects.filter(
            school=school,
            is_active=True
        ).order_by("order")

    
        # ==========================================
        # === POST ACTIONS ===
        # ==========================================
        if request.method == 'POST':
            action = request.POST.get('action')
    
            # ✅ NEW: RENAME PORTS FROM HTML FORM
            if action == 'rename_port':
                key = request.POST.get('port_key', '').strip()
                label = request.POST.get('port_label', '').strip()
            
                if not key or not label:
                    messages.error(request, "❌ Port key / label manquant.")
                else:
                    updated = SchoolPort.objects.filter(
                        school=school,
                        key=key
                    ).update(label=label[:50])
            
                    if updated:
                        messages.success(request, f"✅ Porte '{key}' renommée en '{label}'.")
                    else:
                        messages.error(request, f"❌ Porte '{key}' introuvable.")
            
                return redirect(request.path)   # ✅ يرجّع نفس الصفحة بلا ما تكرر POST
    
            elif action == 'change_login_password':
                new_pass = request.POST.get('new_password')
                confirm_pass = request.POST.get('confirm_password')
    
                if new_pass and confirm_pass:
                    if new_pass == confirm_pass:
                        if len(new_pass) < 6:
                            messages.warning(request, "⚠️ Le mot de passe doit contenir au moins 6 caractères.")
                        else:
                            u = request.user
                            u.set_password(new_pass)
                            u.save()
                            # Important: Bach maydeconnectach l'user
                            update_session_auth_hash(request, u) 
                            messages.success(request, "🔐 Mot de passe de connexion modifié avec succès !")
                    else:
                        messages.error(request, "❌ Les mots de passe ne correspondent pas.")
                else:
                    messages.error(request, "❌ Veuillez remplir les champs.")
    
            # A. UPDATE LOCATION & PIN
            if action == 'update_location':
                new_radius = request.POST.get('radius')
                new_lat = request.POST.get('latitude')
                new_lng = request.POST.get('longitude')
                new_pin = request.POST.get('security_pin')
                is_active = request.POST.get('active_admin_scan')
    
                if new_radius:
                    school.radius = int(new_radius)
    
                if new_lat and new_lng:
                    try:
                        school.latitude = float(new_lat.replace(',', '.'))
                        school.longitude = float(new_lng.replace(',', '.'))
                    except:
                        pass
    
                if new_pin:
                    clean_pin = str(new_pin).strip()
                    if len(clean_pin) == 4 and clean_pin.isdigit():
                        school.security_pin = clean_pin
    
                try:
                    school.active_admin_scan = is_active
                    school.save()
                except:
                    school.active_admin_scan = False
                    school.save()
    
                messages.success(request, "📍 Configuration mise à jour !")
    
            # B. IMPORT CSV
            elif action == 'upload_csv':
                try:
                    csv_file = request.FILES.get('file')
                    if not csv_file.name.endswith('.csv'):
                        messages.error(request, "Le fichier doit être un CSV (.csv)")
                    else:
                        file_data = csv_file.read()
                        try:
                            decoded_file = file_data.decode('utf-8-sig')
                        except UnicodeDecodeError:
                            decoded_file = file_data.decode('latin-1')
    
                        io_string = io.StringIO(decoded_file)
                        first_line = io_string.readline()
                        separator = ';' if ';' in first_line else ','
                        io_string.seek(0)
                        next(io_string, None)
    
                        csv_reader = csv.reader(io_string, delimiter=separator)
                        count = 0
                        for row in csv_reader:
                            try:
                                if len(row) < 7:
                                    continue
                                Student.objects.create(
                                    code_id=row[0].strip(),
                                    class_name=row[1].strip(),
                                    group=row[2].strip(),
                                    last_name=row[3].strip(),
                                    first_name=row[4].strip(),
                                    cndp=row[5].strip(),
                                    badge_number=row[6].strip(),
                                    school=school,
                                )
                                count += 1
                            except:
                                continue
                        messages.success(request, f"✅ {count} élèves ajoutés !")
                except Exception as e:
                    messages.error(request, f"Erreur : {str(e)}")
    
            # B2. IMPORT XLSX
            elif action == 'upload_xlsx':
                try:
                    excel_file = request.FILES.get('file')
    
                    if not excel_file.name.endswith('.xlsx'):
                        messages.error(request, "Le fichier doit être un Excel (.xlsx)")
                    else:
                        try:
                            wb = openpyxl.load_workbook(excel_file)
                            ws = wb.active
    
                            count = 0
                            for row in ws.iter_rows(min_row=2, values_only=True):
                                try:
                                    if not row or len(row) < 7:
                                        continue
    
                                    Student.objects.create(
                                        code_id=str(row[0]).strip() if row[0] else '',
                                        class_name=str(row[1]).strip() if row[1] else '',
                                        group=str(row[2]).strip() if row[2] else '',
                                        last_name=str(row[3]).strip() if row[3] else '',
                                        first_name=str(row[4]).strip() if row[4] else '',
                                        cndp=str(row[5]).strip() if row[5] else '',
                                        badge_number=str(row[6]).strip() if row[6] else '',
                                        school=school,
                                    )
                                    count += 1
                                except:
                                    continue
    
                            messages.success(request, f"✅ {count} élèves ajoutés via Excel !")
    
                        except Exception as e:
                            messages.error(request, f"Erreur de lecture Excel : {str(e)}")
    
                except Exception as e:
                    messages.error(request, f"Erreur globale : {str(e)}")
    
            # C. EXPORT CSV
            elif action == 'export_csv':
                response = HttpResponse(content_type='text/csv')
                response['Content-Disposition'] = 'attachment; filename="eleves_export.csv"'
                response.write(u'\ufeff'.encode('utf8'))
                writer = csv.writer(response, delimiter=';')
                writer.writerow(['code_id','Nom', 'Prénom', 'Classe','group' ,'Numéro Badge', 'Code Massar','school'])
    
                students_export = Student.objects.filter(school=school).order_by('class_name', 'last_name')
                for s in students_export:
                    writer.writerow([s.code_id, s.last_name, s.first_name, s.class_name, s.group, s.badge_number, s.cndp, s.school])
                return response
            
            elif action == 'export_xlsx':
            
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Liste des Elèves"
            
                # 1. تعريف الألوان والتنسيقات
                header_fill = PatternFill(start_color="FFC107", end_color="FFC107", fill_type="solid") # لون أصفر وورنينج
                header_font = Font(bold=True, color="FFFFFF", size=12)
                center_alignment = Alignment(horizontal="center", vertical="center")
                border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            
                # 2. تحديد الرؤوس
                headers = ['Nom', 'Prénom', 'Classe', 'Group', 'Numéro Badge', 'Code Massar']
                ws.append(headers)
            
                # تنسيق سطر الرؤوس
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center_alignment
                    cell.border = border
            
                # 3. إضافة البيانات
                students_export = Student.objects.filter(school=school).order_by('class_name', 'last_name')
                for s in students_export:
                    ws.append([s.last_name, s.first_name, s.class_name, s.group, s.badge_number, s.cndp])
            
                # 4. تفعيل الحماية وتنسيق الأعمدة
                ws.protection.sheet = True
                ws.protection.password = '123'
                
                unlocked = Protection(locked=False)
                locked = Protection(locked=True)
                badge_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid") # لون رمادي للتمييز
            
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    for cell in row:
                        cell.border = border
                        cell.alignment = Alignment(horizontal="left")
                        
                        if cell.column == 5: # عمود Numéro Badge
                            cell.protection = locked
                            cell.fill = badge_fill # تلوينه بالرمادي ليعرف المستخدم أنه للقراءة فقط
                        else:
                            cell.protection = unlocked
            
                # ضبط عرض الأعمدة تلقائياً ليكون الشكل مرتباً
                column_widths = [20, 20, 15, 10, 15, 20]
                for i, width in enumerate(column_widths, 1):
                    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
            
                response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = f'attachment; filename="Eleves_{school.name}.xlsx"'
                wb.save(response)
                return response
            
            
            elif action == 'upload_xlsxV2':
                try:
                    excel_file = request.FILES.get('file')
                    if not excel_file or not excel_file.name.endswith('.xlsx'):
                        messages.error(request, "Veuillez sélectionner un fichier Excel (.xlsx)")
                    else:
                        wb = openpyxl.load_workbook(excel_file)
                        ws = wb.active
                        
                        count_updated = 0
                        not_found_badges = [] # قائمة لتخزين البادجات غير الموجودة
            
                        # نمر على الصفوف (البيانات تبدأ من الصف 2)
                        for row in ws.iter_rows(min_row=2, values_only=True):
                            # الترتيب: 0:Nom, 1:Prénom, 2:Classe, 3:Group, 4:Badge, 5:Massar
                            b_number = str(row[4]).strip() if row[4] else None
                            
                            if not b_number:
                                continue
            
                            # البحث عن التلميذ
                            student = Student.objects.filter(school=school, badge_number=b_number).first()
            
                            if student:
                                # تحديث البيانات
                                student.last_name = str(row[0]).strip() if row[0] else student.last_name
                                student.first_name = str(row[1]).strip() if row[1] else student.first_name 
                                student.class_name = str(row[2]).strip() if row[2] else student.class_name
                                student.group = str(row[3]).strip() if row[3] else student.group
                                student.cndp = str(row[5]).strip() if row[5] else student.cndp
                                student.save()
                                count_updated += 1
                            else:
                                # إذا لم يوجد، نضيف رقم البادج للقائمة لإعلام المستخدم
                                not_found_badges.append(b_number)
            
                        # رسالة النجاح
                        #messages.success(request, f"✅ تم تحديث {count_updated} تلميذ بنجاح.")
                        
                        
                        messages.success(request, f"✅{count_updated} élèves ont été mis à jour avec succès.")
                        # رسالة التنبيه في حال وجود بادجات مفقودة
                        if not_found_badges:
                            badges_str = ", ".join(not_found_badges)
                            #messages.warning(request, f"⚠️ البادجات التالية غير موجودة في قاعدة البيانات ولم يتم تحديثها: {badges_str}")
                            messages.warning(request, f"⚠️ Les badges suivants ne figurent pas dans la base de données et n'ont pas été mis à jour:{badges_str}")
                except Exception as e:
                    messages.error(request, f"Erreur lors de l'import : {str(e)}")
            
            # D. CLEAR LIST
            elif action == 'clear_list':
                Student.objects.filter(school=school).delete()
                messages.warning(request, "🗑️ Liste vidée.")
                
            
            elif action == 'create_reserved':
                try:
                    # --- A. حساب NUMÉRO DE BADGE (Max + 1) ---
                    # 1. نجلب كل أرقام البادجات لهاته المدرسة
                    existing_badges = Student.objects.filter(school=school).values_list('badge_number', flat=True)
                    
                    max_badge = 0
                    for b in existing_badges:
                        # نتأكد أن البادج رقم فقط (نتجاهل RESERVE أو A12)
                        if b and str(b).isdigit():
                            val = int(b)
                            if val > max_badge:
                                max_badge = val
                    
                    # البادج الجديد هو الأكبر + 1
                    next_badge_number = str(max_badge + 1)
    
    
                    # --- B. حساب CODE ID (YYYY YY ID + 001) ---
                    
                    yyyy = datetime.now().year
                    #yyyy = now.strftime("%y")       # "25"
                    school_key = str(school.id)   # "1"
    
                    # الصيغة: 25 + 25 + 1 = 25251
                    prefix_str = f"{yyyy}{yyyy}{school_key}" 
                    
                    # نضرب في 1000 لنترك مساحة للعداد
                    # 25251000
                    base_calculation = int(prefix_str) * 1000
    
                    counter = 1
                    while True:
                        # 25251000 + 1 = 25251001
                        final_result = base_calculation + counter
                        candidate_code = str(final_result)
                        
                        if not Student.objects.filter(code_id=candidate_code).exists():
                            break 
                        counter += 1
    
                    # 4. إنشاء التلميذ
                    Student.objects.create(
                        school=school,
                        first_name="RESERVE",
                        last_name="RESERVE",
                        class_name="RESERVE",
                        group ="RESERVE",
                        # ✅ هنا كنحطو الرقم التسلسلي الجديد (مثلاً 150)
                        badge_number=next_badge_number, 
                        
                        cndp="0000",
                        code_id=candidate_code 
                    )
    
                    messages.success(request, f"✅ Réservé : Badge N° {next_badge_number} (ID: {candidate_code})")
    
                except Exception as e:
                    messages.error(request, f"Erreur Création : {str(e)}")
           
            elif action == 'create_Mreserved':
                try:
                    # 1. جلب العدد المطلوب
                    try:
                        count_to_create = int(request.POST.get('count', 1))
                    except ValueError:
                        count_to_create = 1
            
                    # السنة بـ 2 أرقام (مثلاً 26)
                    yy_short = datetime.now().strftime('%y') 
                    school_key = str(school.api_key)
                    
                    # الـ Prefix دابا غيكون فيه غير 26 + مفتاح المدرسة
                    badge_prefix = f"{yy_short}{school_key}"
                    
                    # --- تحسين الأداء (Optimization) ---
                    existing_badges = set(Student.objects.filter(
                        school=school, 
                        badge_number__startswith=badge_prefix
                    ).values_list('badge_number', flat=True))
                    
                    existing_codes = set(Student.objects.filter(
                        school=school
                    ).values_list('code_id', flat=True))
            
                    current_seq = 1
                    created_count = 0
            
                    # --- حلقة التكرار ---
                    for _ in range(count_to_create):
                        
                        # 1. البحث عن Badge Number متاح
                        while True:
                            # zfill(2) هي اللي كدير لينا 01, 02 عوض 1, 2
                            seq_formatted = str(current_seq).zfill(2)
                            candidate_badge = f"{badge_prefix}{seq_formatted}"
                            
                            if candidate_badge not in existing_badges:
                                break
                            current_seq += 1
                        
                        # 2. لوجيك الـ Code ID (ديالك)
                        # كنستعملو السنة كاملة 2026 في الحساب باش يبقى الرقم فريد
                        yyyy_full = datetime.now().year
                        codage = yyyy_full * current_seq
                        prefix_str = f"{yyyy_full}{school_key}{codage}"
                        base_calculation = int(prefix_str) * 1000
                        
                        counter = 1
                        while True:
                            candidate_code = str(base_calculation + counter)
                            if candidate_code not in existing_codes:
                                break 
                            counter += 1
            
                        # 3. إنشاء السجل
                        Student.objects.create(
                            school=school,
                            first_name="RESERVE",
                            last_name="RESERVE",
                            class_name="RESERVE",
                            group="RESERVE",
                            badge_number=candidate_badge,
                            cndp="0000",
                            code_id=candidate_code 
                        )
                        
                        # تحديث الـ Sets
                        existing_badges.add(candidate_badge)
                        existing_codes.add(candidate_code)
                        
                        current_seq += 1
                        created_count += 1
            
                    messages.success(request, f"✅ تم بنجاح إنشاء {created_count} سجل احتياطي بالتنسيق الجديد.")
            
                except Exception as e:
                    messages.error(request, f"Erreur Création : {str(e)}")
            
            elif action == 'print_badges':
                response = HttpResponse(content_type='application/pdf')
                response['Content-Disposition'] = 'attachment; filename="badges_securises.pdf"'
    
                p = canvas.Canvas(response, pagesize=A4)
                width, height = A4
                
                margin_x = 1 * cm
                margin_y = 1 * cm
                col_width = (width - 2 * margin_x) / 3
                row_height = (height - 2 * margin_y) / 4
    
                students = Student.objects.filter(school=school).order_by('class_name', 'last_name')
    
                col = 0
                row = 3 
                
                # ✅ لائحة الثوابت (نفسها لي فـ Flutter)
                CONSTANTS = [1313131313, 1234567899, 1234567899, 1234567899, 1234567899, 
                             1234567899, 1234567899]
    
                for s in students:
                    x = margin_x + col * col_width
                    y = margin_y + row * row_height
    
                    # Cadre
                    p.setStrokeColorRGB(0.8, 0.8, 0.8)
                    p.rect(x + 0.2*cm, y + 0.2*cm, col_width - 0.4*cm, row_height - 0.4*cm)
    
                    # 1. Numéro Badge (Clair)
                    p.setFont("Helvetica-Bold", 16)
                    display_badge = s.badge_number if s.badge_number else "---"
                    p.drawCentredString(x + col_width/2, y + row_height - 1.5*cm, f"N° {display_badge}")
    
                    # 2. ✅✅✅ CALCUL DU CRYPTAGE (ENCODAGE)
                    try:
                        # نحولو ID لرقم
                        val = int(s.code_id)
                        
                        # الضرب فـ 13 (10 مرات)
                        val = val * (13**10)
                        
                        # زيادة الثوابت 
                        for num in CONSTANTS:
                            val = val + num
                        
                        encrypted_code = str(val)
                    except:
                        # إلا كان الكود فيه حروف، كنخليوه كما هو
                        encrypted_code = encrypted_code
    
                    # 3. QR Code (فيه الكود المشفر)
                    qr = qrcode.QRCode(box_size=10, border=1)
                    qr.add_data(encrypted_code) # 👈 هنا كنحطو الكود الطويل
                    qr.make(fit=True)
                    img_qr = qr.make_image(fill_color="black", back_color="white")
    
                    img_buffer = io.BytesIO()
                    img_qr.save(img_buffer, format='PNG')
                    img_buffer.seek(0)
                    
                    qr_size = 4.5 * cm 
                    qr_y = y + (row_height - qr_size) / 2 - 0.5 * cm 
                    p.drawImage(ImageReader(img_buffer), x + col_width/2 - qr_size/2, qr_y, width=qr_size, height=qr_size)
    
                    col += 1
                    if col >= 3:
                        col = 0
                        row -= 1
                        if row < 0: 
                            p.showPage()
                            row = 3
                
                p.save()
                return response
            
            # ... (dakchi li 9bel)
    
            elif action == 'print_badges_filter':
                # 1. Recuperation des filtres depuis le POST (Hidden inputs)
                filter_class = request.POST.get('filter_class', '').strip()
                filter_q = request.POST.get('filter_q', '').strip()
    
                # 2. Base Query
                students = Student.objects.filter(school=school)
    
                # 3. Appliquer le filtre Classe si existe
                if filter_class:
                    students = students.filter(class_name=filter_class)
    
                # 4. Appliquer le filtre Recherche (kif ma derti f GET)
                if filter_q:
                    students = students.filter(
                        Q(last_name__icontains=filter_q) | 
                        Q(first_name__icontains=filter_q) | 
                        Q(code_id__icontains=filter_q) |
                        Q(badge_number__icontains=filter_q)
                    )
    
                # 5. Ordre
                students = students.order_by('class_name', 'last_name')
    
                # --- GENERATION PDF (Code dyalk l9dim b9a howa howa) ---
                response = HttpResponse(content_type='application/pdf')
                response['Content-Disposition'] = 'attachment; filename="badges_securises.pdf"'
    
                p = canvas.Canvas(response, pagesize=A4)
                width, height = A4
                
                margin_x = 1 * cm
                margin_y = 1 * cm
                col_width = (width - 2 * margin_x) / 3
                row_height = (height - 2 * margin_y) / 4
    
                col = 0
                row = 3 
                
                CONSTANTS = [1313131313, 1234567899, 1234567899, 1234567899, 1234567899, 
                             1234567899, 1234567899]
    
                # Vérification : Ila makayninach talamid
                if not students.exists():
                    messages.warning(request, "Aucun élève à imprimer avec ces filtres.")
                    return redirect('manage_students')
    
                for s in students:
                    # ... (Nfs lcode dyal rasm lbadge li kan 3ndk) ...
                    # ... Met lcode l9dim hna ...
                    x = margin_x + col * col_width
                    y = margin_y + row * row_height
    
                    p.setStrokeColorRGB(0.8, 0.8, 0.8)
                    p.rect(x + 0.2*cm, y + 0.2*cm, col_width - 0.4*cm, row_height - 0.4*cm)
    
                    p.setFont("Helvetica-Bold", 16)
                    display_badge = s.badge_number if s.badge_number else "---"
                    p.drawCentredString(x + col_width/2, y + row_height - 1.5*cm, f"N° {display_badge}")
    
                    try:
                        val = int(s.code_id)
                        val = val * (13**10)
                        for num in CONSTANTS:
                            val = val + num
                        encrypted_code = str(val)
                    except:
                        encrypted_code = s.code_id # Fallback
    
                    qr = qrcode.QRCode(box_size=10, border=1)
                    qr.add_data(encrypted_code)
                    qr.make(fit=True)
                    img_qr = qr.make_image(fill_color="black", back_color="white")
    
                    img_buffer = io.BytesIO()
                    img_qr.save(img_buffer, format='PNG')
                    img_buffer.seek(0)
                    
                    qr_size = 4.5 * cm 
                    qr_y = y + (row_height - qr_size) / 2 - 0.5 * cm 
                    p.drawImage(ImageReader(img_buffer), x + col_width/2 - qr_size/2, qr_y, width=qr_size, height=qr_size)
    
                    col += 1
                    if col >= 3:
                        col = 0
                        row -= 1
                        if row < 0: 
                            p.showPage()
                            row = 3 
                
                p.save()
                return response
            
            
            elif action == 'download_zip_badgesbkp':
                # 1. Recuperation des filtres (nfs l-manti9)
                filter_class = request.POST.get('filter_class', '').strip()
                filter_q = request.POST.get('filter_q', '').strip()
            
                students = Student.objects.filter(school=school)
                if filter_class:
                    students = students.filter(class_name=filter_class)
                if filter_q: 
                    students = students.filter(
                        Q(last_name__icontains=filter_q) | 
                        Q(first_name__icontains=filter_q) | 
                        Q(code_id__icontains=filter_q) |
                        Q(badge_number__icontains=filter_q)
                    )
            
                if not students.exists():
                    messages.warning(request, "Aucun élève trouvé.")
                    return redirect('manage_students')
            
                # 2. Khl9 l-buffer dyal l-ZIP
                zip_buffer = io.BytesIO()
            
                with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                    CONSTANTS = [1313131313, 1234567899, 1234567899, 1234567899, 1234567899, 1234567899, 1234567899]
            
                    for s in students:
                        # Khl9 buffer l kol PDF dyal tilmid
                        pdf_buffer = io.BytesIO()
                        # Ghadi nkhdmo b size sghir (masalan 10cm x 10cm) hit kolla PDF fih QR wahd
                        # Ila bghiti A4 khliha A4
                        p = canvas.Canvas(pdf_buffer, pagesize=(10*cm, 10*cm))
                        
                        # --- CALCUL DU CRYPTAGE ---
                        try:
                            val = int(s.code_id)
                            val = val * (13**10)
                            for num in CONSTANTS:
                                val = val + num
                            encrypted_code = str(val)
                        except:
                            encrypted_code = s.code_id
            
                        # --- GENERATION QR ---
                        qr = qrcode.QRCode(box_size=10, border=1)
                        qr.add_data(encrypted_code)
                        qr.make(fit=True)
                        img_qr = qr.make_image(fill_color="black", back_color="white")
            
                        qr_temp_buffer = io.BytesIO()
                        img_qr.save(qr_temp_buffer, format='PNG')
                        qr_temp_buffer.seek(0)
            
                        # --- RASM F PDF ---
                        p.setFont("Helvetica-Bold", 12)
                        display_name = f"SCAN ME "
                        p.drawCentredString(5*cm, 8.5*cm, display_name)
                        p.drawCentredString(5*cm, 1.5*cm, f"N° {s.badge_number or '---'}")
                        
                        p.drawImage(ImageReader(qr_temp_buffer), 2.5*cm, 2.5*cm, width=5*cm, height=5*cm)
                        
                        p.showPage()
                        p.save()
            
                        # 3. Zid l-PDF l l-ZIP
                        pdf_buffer.seek(0)
                        file_name = f"badge_{s.badge_number or s.id}_{s.last_name}.pdf"
                        zip_file.writestr(file_name, pdf_buffer.read())
            
                # 4. Erja3 l-ZIP kaml l-user
                zip_buffer.seek(0)
                response = HttpResponse(zip_buffer.read(), content_type='application/zip')
                response['Content-Disposition'] = 'attachment; filename="badges_eleves.zip"'
                return response
            
            elif action == 'download_zip_badges':
                # 1. Recuperation des filtres (nfs l-manti9)
                filter_class = request.POST.get('filter_class', '').strip()
                filter_q = request.POST.get('filter_q', '').strip()
            
                students = Student.objects.filter(school=school)
                if filter_class:
                    students = students.filter(class_name=filter_class)
                if filter_q: 
                    students = students.filter(
                        Q(last_name__icontains=filter_q) | 
                        Q(first_name__icontains=filter_q) | 
                        Q(code_id__icontains=filter_q) |
                        Q(badge_number__icontains=filter_q)
                    )
            
                if not students.exists():
                    messages.warning(request, "Aucun élève trouvé.")
                    return redirect('manage_students')
            
                # 2. Khl9 l-buffer dyal l-ZIP
                zip_buffer = io.BytesIO()
            
                with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                    
            
                    for s in students:
                        # Khl9 buffer l kol PDF dyal tilmid
                        pdf_buffer = io.BytesIO()
                        # Ghadi nkhdmo b size sghir (masalan 10cm x 10cm) hit kolla PDF fih QR wahd
                        # Ila bghiti A4 khliha A4
                        p = canvas.Canvas(pdf_buffer, pagesize=(10*cm, 10*cm))
                        
                        # --- CALCUL DU CRYPTAGE ---
                        try:
                            val = int(s.code_id)
                        
                            
                            encrypted_code = str(val)
                        except:
                            encrypted_code = s.code_id
            
                        # --- GENERATION QR ---
                        qr = qrcode.QRCode(box_size=10, border=1)
                        qr.add_data(encrypted_code)
                        qr.make(fit=True)
                        img_qr = qr.make_image(fill_color="black", back_color="white")
            
                        qr_temp_buffer = io.BytesIO()
                        img_qr.save(qr_temp_buffer, format='PNG')
                        qr_temp_buffer.seek(0)
            
                        # --- RASM F PDF ---
                        p.setFont("Helvetica-Bold", 12)
                        display_name = f"SCAN ME "
                        p.drawCentredString(5*cm, 8.5*cm, display_name)
                        p.drawCentredString(5*cm, 1.5*cm, f"N° {s.badge_number or '---'}")
                        
                        p.drawImage(ImageReader(qr_temp_buffer), 2.5*cm, 2.5*cm, width=5*cm, height=5*cm)
                        
                        p.showPage()
                        p.save()
            
                        # 3. Zid l-PDF l l-ZIP
                        pdf_buffer.seek(0)
                        file_name = f"badge_{s.badge_number or s.id}_{s.last_name}.pdf"
                        zip_file.writestr(file_name, pdf_buffer.read())
            
                # 4. Erja3 l-ZIP kaml l-user
                zip_buffer.seek(0)
                response = HttpResponse(zip_buffer.read(), content_type='application/zip')
                response['Content-Disposition'] = 'attachment; filename="badges_eleves.zip"'
                return response
            
            
            elif action == 'delete_student':
                try:
                    student_id = request.POST.get('student_id')
                    
                    # البحث عن التلميذ (مع التأكد أنه تابع لنفس المدرسة للحماية)
                    student = Student.objects.get(id=student_id, school=school)
                    
                    # الحذف
                    name = f"{student.first_name} {student.last_name}"
                    student.delete()
                    
                    messages.success(request, f"🗑️ L'élève {name} a été supprimé.")
                    
                except Student.DoesNotExist:
                    messages.error(request, "❌ Élève introuvable.")
                except Exception as e:
                    messages.error(request, f"Erreur : {str(e)}")
           
            elif action == 'add_garde_category':
                cat_name = request.POST.get('category_name')
                if cat_name:
                    GardeCategory.objects.create(school=school, name=cat_name)
                    messages.success(request, f"✅ Catégorie '{cat_name}' créée !")
    
            # B. AJOUTER/MODIFIER CLASSE AVEC GARDE
            elif action == 'add_class':
                class_name = request.POST.get('new_class_name', '').strip().upper()
                garde_id = request.POST.get('garde_id') # <-- Hada jdid
                
                if class_name:
                    # Njibo Garde object ila kan choisi
                    selected_garde = None
                    if garde_id:
                        try:
                            selected_garde = GardeCategory.objects.get(id=garde_id, school=school)
                        except:
                            pass
    
                    # N9albo wach classe kayna wla la (Update or Create)
                    obj, created = SchoolClass.objects.update_or_create(
                        school=school, 
                        name=class_name,
                        defaults={'garde_category': selected_garde} # <-- Hna kanls9o l garde
                    )
                    
                    if created:
                        messages.success(request, f"✅ Classe {class_name} ajoutée avec succès !")
                    else:
                        messages.info(request, f"🔄 Classe {class_name} mise à jour.")
                    show_config_modal = True
                
               
           
           # ... (dakchi li kan 9bel: add_garde_category, add_class...)
    
            # 3. MODIFIER UNE GARDE
            elif action == 'edit_garde':
                garde_id = request.POST.get('garde_id')
                new_name = request.POST.get('new_name')
                if garde_id and new_name:
                    try:
                        garde = GardeCategory.objects.get(id=garde_id, school=school)
                        garde.name = new_name
                        garde.save()
                        messages.success(request, f"✅ Garde renommée en '{new_name}'.")
                    except GardeCategory.DoesNotExist:
                        messages.error(request, "Erreur: Garde introuvable.")
                show_config_modal = True
    
            # 4. SUPPRIMER UNE GARDE
            elif action == 'delete_garde':
                garde_id = request.POST.get('garde_id')
                if garde_id:
                    try:
                        garde = GardeCategory.objects.get(id=garde_id, school=school)
                        garde_name = garde.name
                        garde.delete()
                        # Melli katmsse7, les classes kaywlliw NULL automatiquement hit drna on_delete=SET_NULL
                        messages.warning(request, f"🗑️ La catégorie '{garde_name}' a été supprimée.")
                    except:
                        messages.error(request, "Erreur lors de la suppression.")
                show_config_modal = True
    
            # 5. ANNULER L'AFFECTATION (Détacher une classe)
            elif action == 'remove_class_garde':
                class_id = request.POST.get('class_id')
                if class_id:
                    try:
                        cls = SchoolClass.objects.get(id=class_id, school=school)
                        cls.garde_category = None # Hna kanraj3oha khawya
                        cls.save()
                        messages.info(request, f"🔗 Garde retirée pour la classe {cls.name}.")
                    except:
                        messages.error(request, "Erreur: Classe introuvable.")
                        
                show_config_modal = True
           
    
            # 6. CRÉER UNE CLASSE SEULEMENT (Sans Garde)
            elif action == 'create_simple_class':
                class_name = request.POST.get('class_name', '').strip().upper()
                if class_name:
                    # get_or_create: Ila kant kayjibha, ila makantch kaycréerha
                    obj, created = SchoolClass.objects.get_or_create(school=school, name=class_name)
                    
                    if created:
                        messages.success(request, f"✅ Classe '{class_name}' créée.")
                    else:
                        messages.warning(request, f"⚠️ La classe '{class_name}' existe déjà.")
                
                show_config_modal = True # Bach tb9a lmodal m7loula
            
            # F west manage_students (m3a actions POST)
     
            elif action == 'assign_class_garde':
                try:
                    class_id = request.POST.get('class_id') # ID dyal classe li bghina nzido
                    garde_id = request.POST.get('garde_id') # ID dyal Garde fin bghina nkhchiwha
                    
                    school_class = get_object_or_404(SchoolClass, id=class_id, school=school)
                    garde = get_object_or_404(GardeCategory, id=garde_id, school=school)
                    
                    school_class.garde_category = garde
                    school_class.save()
                    
                    messages.success(request, f"Classe {school_class.name} ajoutée à {garde.name}")
                except Exception as e:
                    messages.error(request, f"Erreur d'affectation : {str(e)}") 
        
            elif action == 'remove_class_garde':
                try:
                    class_id = request.POST.get('class_id')
                    school_class = get_object_or_404(SchoolClass, id=class_id, school=school)
                    
                    old_garde = school_class.garde_category.name if school_class.garde_category else ""
                    school_class.garde_category = None
                    school_class.save()
                    
                    messages.success(request, f"Classe {school_class.name} retirée de {old_garde}")
                except Exception as e:
                    messages.error(request, f"Erreur : {str(e)}")
            
            elif action == 'delete_selected':
                selected_ids = request.POST.getlist('student_ids')
                confirm_pass = request.POST.get('confirm_pass_input') # تأكد أن هذا الاسم يطابق الـ JS
            
                if confirm_pass == "0000":
                    if selected_ids:
                        # تنفيذ الحذف
                        deleted_count = Student.objects.filter(id__in=selected_ids, school=school).delete()[0]
                        messages.success(request, f"✅ {deleted_count} élève(s) supprimé(s) avec succès.")
                    else:
                        messages.warning(request, "⚠️ Veuillez sélectionner au moins un élève.")
                else:
                    # هذه الرسالة التي ظهرت لك تعني أن confirm_pass لم تكن "000"
                    messages.error(request, "❌ Mot de passe de confirmation incorrect. Suppression annulée.")
                    
            
            
            
                
            if action in ['add_garde_category', 'add_configured_class', 'edit_garde', 'delete_garde', 'remove_class_garde']:
                
                # ... (Lcode li kan 3ndk hna dyal traitemement khelih kif ma howa) ...
                
                # Zid hadi f kol block, awla f lakher d had actions:
                show_config_modal = True
           
            
    
            # ✅ باقي actions ديالك (create_reserved, print_badges, zip, delete_student, garde, class...)
            # خليتهم كيف ما هما، ما بدلت فيهم والو (نفس الكود ديالك)
    
            # ... continue your existing blocks here unchanged ...
    
    
        # ==========================================
        # === GET DISPLAY ===
        # ==========================================
    
        students_list = Student.objects.filter(school=school)\
            .annotate(cleaned_class=Trim(Lower('class_name')))\
            .order_by('cleaned_class', 'last_name')
    
        # a. classes من students
        classes_from_students = set(
            Student.objects.filter(school=school)
            .exclude(class_name__isnull=True)
            .exclude(class_name__exact='')
            .values_list('class_name', flat=True)
        )
    
        # b. classes من config
        classes_from_config = set(
            SchoolClass.objects.filter(school=school)
            .values_list('name', flat=True)
        )
    
        # c. union
        all_classes_set = classes_from_students.union(classes_from_config)
        all_classes = list(all_classes_set)
    
        saved_order_str = school.class_order
        if saved_order_str:
            saved_order = [x.strip() for x in saved_order_str.split(',') if x.strip()]
            order_map = {name: i for i, name in enumerate(saved_order)}
            all_classes.sort(key=lambda x: order_map.get(x, 999))
        else:
            all_classes.sort()
    
        garde_categories = GardeCategory.objects.filter(school=school)
        configured_classes = SchoolClass.objects.filter(school=school).select_related('garde_category').order_by('name')
        student_classes_list = all_classes
    
        search_query = request.GET.get('q', '').strip()
        selected_class = request.GET.get('classe', '').strip()
    
        if search_query:
            students_list = students_list.filter(
                Q(last_name__icontains=search_query) |
                Q(first_name__icontains=search_query) |
                Q(code_id__icontains=search_query) |
                Q(badge_number__icontains=search_query)
            )
    
        if selected_class:
            students_list = students_list.filter(class_name=selected_class)
    
        per_page = request.GET.get('per_page', '100')
        try:
            per_page = int(per_page)
        except:
            per_page = 100
    
        paginator = Paginator(students_list, per_page)
        page_number = request.GET.get('page')
        students = paginator.get_page(page_number)
    
        # ✅ IMPORTANT: ports خاصها تتبعث هنا للـ template
        return render(request, 'manage_admin.html', {
            'school': school,
            'students': students,
            'per_page': per_page,
            'all_classes': all_classes,
            'search_query': search_query,
            'selected_class': selected_class,
            'garde_categories': garde_categories,
            'configured_classes': configured_classes,
            'student_classes_list': student_classes_list,
            'show_config_modal': show_config_modal,
    
            # ✅ NEW
            'ports': ports,
        })

    except:
        return render(request, 'manage_admin.html')

   
@login_required
def manage_students_NC(request):
    try:
        school = request.user.profile.school
    except:
        return render(request, 'error.html', {'message': "Pas d'école assignée."})

    # === POST ACTIONS ===
    if request.method == 'POST':
        action = request.POST.get('action')

        # A. UPDATE LOCATION & PIN
        if action == 'update_locationNC':
            new_radius = request.POST.get('radius') 
            new_lat = request.POST.get('latitude')
            new_lng = request.POST.get('longitude')
            new_pin = request.POST.get('security_pin')
            is_active = request.POST.get('active_admin_scan') 
            

            if new_radius: school.radius = int(new_radius)
            if new_lat and new_lng:
                try:
                    school.latitude = float(new_lat.replace(',', '.'))
                    school.longitude = float(new_lng.replace(',', '.'))
                except: pass
            
            if new_pin:
                clean_pin = str(new_pin).strip()
                if len(clean_pin) == 4 and clean_pin.isdigit():
                    school.security_pin = clean_pin
            
            
            try:
                
                school.active_admin_scan = is_active
                school.save()
            except:
                school.active_admin_scan = False
                school.save()
                
            messages.success(request, "📍 Configuration mise à jour !")

        # B. IMPORT CSV
        elif action == 'upload_csv':
            try:
                csv_file = request.FILES.get('file')
                if not csv_file.name.endswith('.csv'):
                    messages.error(request, "Le fichier doit être un CSV (.csv)")
                else:
                    file_data = csv_file.read()
                    try:
                        decoded_file = file_data.decode('utf-8-sig')
                    except UnicodeDecodeError:
                        decoded_file = file_data.decode('latin-1')

                    io_string = io.StringIO(decoded_file)
                    first_line = io_string.readline()
                    separator = ';' if ';' in first_line else ','
                    io_string.seek(0)
                    next(io_string, None)

                    csv_reader = csv.reader(io_string, delimiter=separator)
                    count = 0
                    for row in csv_reader:
                        try:
                            if len(row) < 4: continue
                            Student.objects.create(
                                code_id=row[0].strip(),
                                class_name=row[1].strip(),
                                group =row[2].strip(),
                                last_name=row[3].strip(),
                                first_name=row[4].strip(),
                                cndp  =row[5].strip(),
                                badge_number=row[6].strip(),
                                school=school,
                            )
                            count += 1
                        except: continue
                    messages.success(request, f"✅ {count} élèves ajoutés !")
            except Exception as e:
                messages.error(request, f"Erreur : {str(e)}")

        # C. EXPORT CSV
        elif action == 'export_csv':
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = 'attachment; filename="eleves_export.csv"'
            response.write(u'\ufeff'.encode('utf8'))
            writer = csv.writer(response, delimiter=';')
            writer.writerow(['Nom', 'Prénom', 'Classe', 'Numéro Badge', 'Code Massar'])
            students_export = Student.objects.filter(school=school).order_by('class_name', 'last_name')
            
            for s in students_export:
                writer.writerow([s.last_name, s.first_name, s.class_name, s.badge_number, s.cndp])
            return response

        # D. CLEAR LIST
        elif action == 'clear_list':
            Student.objects.filter(school=school).delete()
            messages.warning(request, "🗑️ Liste vidée.")

        elif action == 'create_reserved':
            try:
                # --- A. حساب NUMÉRO DE BADGE (Max + 1) ---
                # 1. نجلب كل أرقام البادجات لهاته المدرسة
                existing_badges = Student.objects.filter(school=school).values_list('badge_number', flat=True)
                
                max_badge = 0
                for b in existing_badges:
                    # نتأكد أن البادج رقم فقط (نتجاهل RESERVE أو A12)
                    if b and str(b).isdigit():
                        val = int(b)
                        if val > max_badge:
                            max_badge = val
                
                # البادج الجديد هو الأكبر + 1
                next_badge_number = str(max_badge + 1)


                # --- B. حساب CODE ID (YYYY YY ID + 001) ---
                
                yyyy = datetime.now().year
                #yyyy = now.strftime("%y")       # "25"
                school_key = str(school.id)   # "1"

                # الصيغة: 25 + 25 + 1 = 25251
                prefix_str = f"{yyyy}{yyyy}{school_key}" 
                
                # نضرب في 1000 لنترك مساحة للعداد
                # 25251000
                base_calculation = int(prefix_str) * 1000

                counter = 1
                while True:
                    # 25251000 + 1 = 25251001
                    final_result = base_calculation + counter
                    candidate_code = str(final_result)
                    
                    if not Student.objects.filter(code_id=candidate_code).exists():
                        break 
                    counter += 1

                # 4. إنشاء التلميذ
                Student.objects.create(
                    school=school,
                    first_name="RESERVE",
                    last_name="RESERVE",
                    class_name="RESERVE",
                    
                    # ✅ هنا كنحطو الرقم التسلسلي الجديد (مثلاً 150)
                    badge_number=next_badge_number, 
                    
                    cndp="RESERVE",
                    code_id=candidate_code 
                )

                messages.success(request, f"✅ Réservé : Badge N° {next_badge_number} (ID: {candidate_code})")

            except Exception as e:
                messages.error(request, f"Erreur Création : {str(e)}")
       
        elif action == 'print_badges':
            response = HttpResponse(content_type='application/pdf')
            response['Content-Disposition'] = 'attachment; filename="badges_securises.pdf"'

            p = canvas.Canvas(response, pagesize=A4)
            width, height = A4
            
            margin_x = 1 * cm
            margin_y = 1 * cm
            col_width = (width - 2 * margin_x) / 3
            row_height = (height - 2 * margin_y) / 4

            students = Student.objects.filter(school=school).order_by('class_name', 'last_name')

            col = 0
            row = 3 
            
            # ✅ لائحة الثوابت (نفسها لي فـ Flutter)
            CONSTANTS = [1313131313, 1234567899, 1234567899, 1234567899, 1234567899, 
                         1234567899, 1234567899]

            for s in students:
                x = margin_x + col * col_width
                y = margin_y + row * row_height

                # Cadre
                p.setStrokeColorRGB(0.8, 0.8, 0.8)
                p.rect(x + 0.2*cm, y + 0.2*cm, col_width - 0.4*cm, row_height - 0.4*cm)

                # 1. Numéro Badge (Clair)
                p.setFont("Helvetica-Bold", 16)
                display_badge = s.badge_number if s.badge_number else "---"
                p.drawCentredString(x + col_width/2, y + row_height - 1.5*cm, f"N° {display_badge}")

                # 2. ✅✅✅ CALCUL DU CRYPTAGE (ENCODAGE)
                try:
                    # نحولو ID لرقم
                    val = int(s.code_id)
                    
                    
                    
                    encrypted_code = str(val)
                except:
                    # إلا كان الكود فيه حروف، كنخليوه كما هو
                    encrypted_code = encrypted_code

                # 3. QR Code (فيه الكود المشفر)
                qr = qrcode.QRCode(box_size=10, border=1)
                qr.add_data(encrypted_code) # 👈 هنا كنحطو الكود الطويل
                qr.make(fit=True)
                img_qr = qr.make_image(fill_color="black", back_color="white")

                img_buffer = io.BytesIO()
                img_qr.save(img_buffer, format='PNG')
                img_buffer.seek(0)
                
                qr_size = 4.5 * cm 
                qr_y = y + (row_height - qr_size) / 2 - 0.5 * cm 
                p.drawImage(ImageReader(img_buffer), x + col_width/2 - qr_size/2, qr_y, width=qr_size, height=qr_size)

                col += 1
                if col >= 3:
                    col = 0
                    row -= 1
                    if row < 0: 
                        p.showPage()
                        row = 3
            
            p.save()
            return response
        
        elif action == 'delete_student':
            try:
                student_id = request.POST.get('student_id')
                
                # البحث عن التلميذ (مع التأكد أنه تابع لنفس المدرسة للحماية)
                student = Student.objects.get(id=student_id, school=school)
                
                # الحذف
                name = f"{student.first_name} {student.last_name}"
                student.delete()
                
                messages.success(request, f"🗑️ L'élève {name} a été supprimé.")
                
            except Student.DoesNotExist:
                messages.error(request, "❌ Élève introuvable.")
            except Exception as e:
                messages.error(request, f"Erreur : {str(e)}")
       
        return redirect('manage_students_NC')

    # === GET DISPLAY ===
    students_list = Student.objects.filter(school=school)\
    .annotate(cleaned_class=Trim(Lower('class_name')))\
    .order_by('cleaned_class', 'last_name')

    all_classes = Student.objects.filter(school=school)\
        .exclude(class_name__isnull=True).exclude(class_name__exact='')\
        .values_list('class_name', flat=True).distinct().order_by('class_name')
        
    

    search_query = request.GET.get('q', '').strip()
    selected_class = request.GET.get('classe', '').strip()

    if search_query:
        students_list = students_list.filter(
            Q(last_name__icontains=search_query) | 
            Q(first_name__icontains=search_query) | 
            Q(code_id__icontains=search_query) |
            Q(badge_number__icontains=search_query)
        )

    if selected_class:
        students_list = students_list.filter(class_name=selected_class)

    per_page = request.GET.get('per_page', '100')
    try: per_page = int(per_page)
    except: per_page = 100

    paginator = Paginator(students_list, per_page)
    page_number = request.GET.get('page')
    students = paginator.get_page(page_number)

    return render(request, 'manage_students_NC.html', {
        'school': school, 
        'students': students,
        'per_page': per_page,
        'all_classes': all_classes,
        'search_query': search_query,
        'selected_class': selected_class
    })



# ... (تأكد أنك قمت باستيراد qrcode و io)
def generate_student_qr1(request, school_key, student_id):
    try:
        student = Student.objects.get(id=student_id)
        
        # --- 1. LOGIC CRYPTAGE ---
        CONSTANTS = [1313131313, 1234567899, 1234567899, 1234567899, 1234567899, 
                     1234567899, 1234567899]
        
        try:
            # 1. N-testiw wach kyn code_id aslan
            if not student.code_id:
                raise ValueError("Pas de code_id")

            # 2. N7awlo ID l ra9m
            val = int(student.code_id)
            
            # 3. Darb f 13 (10 merrat)
            val = val * (13**10)
            
            # 4. Ziada dyal Constants
            for num in CONSTANTS:
                val = val + num
            
            encrypted_code = str(val)

        except (ValueError, TypeError, OverflowError):
            # ✅ TSHI7: Hna fin kan l-mochkil.
            # Daba ila w93 ghalat, n-akhdo code l-asli d tilmid direct.
            # Ila kan code_id khawi, n-akhdo ID dyal base de données.
            encrypted_code = str(student.code_id) if student.code_id else str(student.id)
        
        # --- 2. GENERATION IMAGE ---
        qr = qrcode.QRCode(box_size=10, border=1)
        qr.add_data(encrypted_code)
        qr.make(fit=True)
        img = qr.make_image(fill_color="black", back_color="white")
        
        # --- 3. REPONSE (TELECHARGEMENT) ---
        response = HttpResponse(content_type="image/png")
        
        # Géri l-header d telechargement
        if request.GET.get('download'):
            # Smiya nqiya bla machakil d l-espaces
            filename = f"Badge_{student.id}.png"
            if student.badge_number:
                filename = f"Badge_{student.badge_number}.png"
                
            response['Content-Disposition'] = f'attachment; filename="{filename}"'

        # Sauvgardi l-image f response
        img.save(response, "PNG")
        return response

    except Student.DoesNotExist:
        return HttpResponse("Élève introuvable", status=404)
    except Exception as e:
        # Hada bach ila w93 chi mochkil akhor, yban lik message f l-fichier
        return HttpResponse(f"Erreur Serveur: {str(e)}", status=500)
  

def generate_student_qr(request, school_key, student_id):
    try:
        student = Student.objects.get(id=student_id)
        
        # --- 1. LOGIC CRYPTAGE ---
        CONSTANTS = [1313131313, 1234567899, 1234567899, 1234567899, 1234567899, 
                     1234567899, 1234567899]
        
        try:
            val = int(student.code_id)
            val = val * (13**10)
            for num in CONSTANTS:
                val = val + num
            encrypted_code = str(val)
        except:
            encrypted_code = str(student.code_id) if student.code_id else str(student.id)
        
        # --- 2. GENERATION QR CODE ---
        # Ba9yin f border=1 bach tb9a l-image sghira bla 7achiya bayda kbira
        qr = qrcode.QRCode(version=3, box_size=60, border=1) 
        qr.add_data(encrypted_code)
        qr.make(fit=True)

        qr_img = qr.make_image(fill_color='black', back_color='white').convert('RGB')

        # --- 3. ZIADA DYAL BADGE NUMBER ---
        badge_text =  f"NR:{student.badge_number}"  if student.badge_number else ""
        
        if badge_text:
            font_size = 200 
            
            try:
                font = ImageFont.truetype("arial.ttf", font_size)
            except IOError:
                try:
                    font = ImageFont.load_default(size=font_size)
                except:
                    font = ImageFont.load_default() 

            draw_temp = ImageDraw.Draw(qr_img)
            text_bbox = draw_temp.textbbox((0, 0), badge_text, font=font)
            text_width = text_bbox[2] - text_bbox[0]
            text_height = text_bbox[3] - text_bbox[1]
            
            # --- 4. HSSAB L-POSITION (CORRECTION) ---
            
            # ✅ Zidna hna: kant 50, rddinaha 150 bach l-ktba mat9ta3ch mn lfo9
            margin_top = 150 
            
            # ✅ Zidna hna: kant 10, rddinaha 30.
            # Haka ghatb9a 9riba bzaf walakin ma-ghatghttich 3la les points noirs.
            margin_middle = 30
            
            margin_bottom = 50

            new_width = max(qr_img.size[0], text_width + 100)
            new_height = margin_top + text_height + margin_middle + qr_img.size[1] + margin_bottom
            
            final_img = Image.new('RGB', (new_width, new_height), 'white')
            
            draw = ImageDraw.Draw(final_img)

            # A. N-ktbo Rr9m
            text_x = (new_width - text_width) // 2 
            text_y = margin_top
            
            # Astuce sghira: kn-n9sso chwiya mn text_bbox[1] bach n-caliw l-font mzyan
            draw.text((text_x, text_y - text_bbox[1]), badge_text, fill="black", font=font)

            # B. N-coller l-QR Code
            qr_x = (new_width - qr_img.size[0]) // 2
            
            # Hna qr_y = fin salat l-ktba + margin_middle
            qr_y = margin_top + text_height + margin_middle
            
            final_img.paste(qr_img, (qr_x, qr_y))
            
        else:
            final_img = qr_img

        # --- 5. RETURN RESPONSE ---
        response = HttpResponse(content_type="image/png")
        if request.GET.get('download'):
            filename = f"Badge_{student.badge_number if student.badge_number else student.id}.png"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'

        final_img.save(response, "PNG")
        return response

    except Student.DoesNotExist:
        return HttpResponse(status=404)
        
def generate_student_qrNC(request, school_key,student_id):
    try:
        
        student = Student.objects.get(id=student_id)
        
        # --- نفس منطق التشفير ---
                     
        
                         
            
        try:
            # نحولو ID لرقم
            val = int(student.code_id)
            
            
            
            encrypted_code = str(val)
        except:
            # إلا كان الكود فيه حروف، كنخليوه كما هو
            encrypted_code = encrypted_code
        
        
            
        # --- إنشاء الصورة ---
        qr = qrcode.QRCode(box_size=10, border=1)
        qr.add_data(encrypted_code)
        qr.make(fit=True)
        img = qr.make_image(fill_color="black", back_color="white")
        

        
        response = HttpResponse(content_type="image/png")
        img.save(response, "PNG")
        return response
        
    except Student.DoesNotExist:
        return HttpResponse(status=404)



# --- 2. EDIT STUDENT (Retour à la même page) ---
@login_required
def edit_student(request, student_id):
    school = request.user.profile.school
    student = get_object_or_404(Student, id=student_id, school=school)
    current_page = request.GET.get('page') or request.POST.get('page') or '1'
    current_per_page = request.GET.get('per_page') or request.POST.get('per_page') or '10'

    if request.method == 'POST':
        form = StudentForm(request.POST, instance=student, school=school)
        if form.is_valid():
            form.save()
            base_url = reverse('manage_students')
            return redirect(f"{base_url}?page={current_page}&per_page={current_per_page}")
    else:
        form = StudentForm(instance=student, school=school)

    return render(request, 'edit_student.html', {
        'form': form, 'student': student,
        'current_page': current_page, 'current_per_page': current_per_page
    })


@login_required
def edit_studentNC(request, student_id):
    school = request.user.profile.school
    student = get_object_or_404(Student, id=student_id, school=school)
    current_page = request.GET.get('page') or request.POST.get('page') or '1' 
    current_per_page = request.GET.get('per_page') or request.POST.get('per_page') or '10'

    if request.method == 'POST':
        form = StudentForm(request.POST, instance=student, school=school)
        if form.is_valid():
            form.save()
            base_url = reverse('manage_students_NC')
            return redirect(f"{base_url}?page={current_page}&per_page={current_per_page}")
    else:
        form = StudentForm(instance=student, school=school)

    return render(request, 'edit_student.html', {
        'form': form, 'student': student,
        'current_page': current_page, 'current_per_page': current_per_page
    })



# (Reste des fonctions API: set_admin_badge, record_scan, etc. ne changent pas)
# 2. تعديل تلميذ
@login_required
def edit_student(request, student_id):
    school = request.user.profile.school
    student = get_object_or_404(Student, id=student_id, school=school)

    # 1. استقبال رقم الصفحة والعدد (من الرابط أو من الفورم)
    # إلا ما كانوش، كنديرو 1 و 10 كقيم افتراضية
    current_page = request.GET.get('page') or request.POST.get('page') or '1'
    current_per_page = request.GET.get('per_page') or request.POST.get('per_page') or '10'

    if request.method == 'POST':
        form = StudentForm(request.POST, instance=student, school=school)
        if form.is_valid():
            form.save()
            
            # 2.  الرجوع لنفس الصفحة بالضبط
            # كنصايبو الرابط الجديد بيدينا
            base_url = reverse('manage_students')
            return redirect(f"{base_url}?page={current_page}&per_page={current_per_page}")
            
    else:
        form = StudentForm(instance=student, school=school)

    return render(request, 'edit_student.html', {
        'form': form, 
        'student': student,
        # 3. نصيفطو الأرقام للصفحة باش تخبيهم ف inputs hidden
        'current_page': current_page,
        'current_per_page': current_per_page
    })
    
    
@api_view(['POST'])
@permission_classes([AllowAny])
def update_school_location(request):
    try:
        school = request.user.profile.school
        
        lat = request.data.get('latitude')
        lng = request.data.get('longitude')
        
        if lat is None or lng is None:
            return Response({"error": "Coordonnées manquantes"}, status=400)

        school.latitude = float(lat)
        school.longitude = float(lng)
        
        #      تصحيح أوتوماتيكي: إلا كان Radius يساوي 0، ردو 100
        if school.radius <= 0:
            school.radius = 100
            
        school.save()

        return Response({"message": "Position mise à jour !"})
    except Exception as e:
        return Response({"error": str(e)}, status=400)
        



@login_required
def school_dashboard(request): # C'est la même fonction que manage_students
    try:
        school = request.user.profile.school
    except:
        return render(request, 'error.html', {'message': "Vous n'avez pas d'école assignée."})

    # ==========================================
    # 1. TRAITEMENT DES ACTIONS (POST)
    # ==========================================
    if request.method == 'POST':
        action = request.POST.get('action')

        # --- A. MISE À JOUR GPS & SÉCURITÉ ---
        if action == 'update_location':
            new_radius = request.POST.get('radius')
            new_lat = request.POST.get('latitude')
            new_lng = request.POST.get('longitude')
            new_pin = request.POST.get('security_pin')
            is_active = request.POST.get('active_admin_scan') 
            # Radius
            if new_radius:
                school.radius = int(new_radius)
            
            # GPS
            if new_lat and new_lng:
                try:
                    school.latitude = float(new_lat.replace(',', '.'))
                    school.longitude = float(new_lng.replace(',', '.'))
                except ValueError:
                    pass

            #  PIN DE SÉCURITÉ
            if new_pin:
                clean_pin = str(new_pin).strip()
                if len(clean_pin) == 4 and clean_pin.isdigit():
                    school.security_pin = clean_pin
                else:
                    messages.warning(request, "️ Le code PIN doit contenir exactement 4 chiffres.")
                    
            school.active_admin_scan = is_active
            school.save()
            messages.success(request, " Configuration (GPS & PIN) mise à jour !")

        # --- B. IMPORTATION CSV ---
        elif action == 'upload_csv':
            try:
                csv_file = request.FILES.get('file')
                if not csv_file.name.endswith('.csv'):
                    messages.error(request, "Le fichier doit être un CSV (.csv)")
                else:
                    file_data = csv_file.read()
                    try:
                        decoded_file = file_data.decode('utf-8-sig')
                    except UnicodeDecodeError:
                        decoded_file = file_data.decode('latin-1')

                    io_string = io.StringIO(decoded_file)
                    first_line = io_string.readline()
                    separator = ';' if ';' in first_line else ','
                    io_string.seek(0)
                    next(io_string, None)

                    csv_reader = csv.reader(io_string, delimiter=separator)
                    count = 0
                    for row in csv_reader:
                        try:
                            if len(row) < 4: continue
                            Student.objects.create(
                                school=school,
                                last_name=row[0].strip(),
                                first_name=row[1].strip(),
                                class_name=row[2].strip(),
                                code_id=row[3].strip(),
                                badge_number=row[3].strip()
                            )
                            count += 1
                        except:
                            continue
                    
                    if count > 0:
                        messages.success(request, f" {count} élèves ajoutés !")
                    else:
                        messages.warning(request, "Aucun élève ajouté.")

            except Exception as e:
                messages.error(request, f"Erreur CSV : {str(e)}")

        # --- C. VIDER LA LISTE ---
        elif action == 'clear_list':
            Student.objects.filter(school=school).delete()
            messages.warning(request, "🗑️ Liste vidée.")

        return redirect('school_dashboard')

    # ==========================================
    # 2. PRÉPARATION DE L'AFFICHAGE (GET)
    # ==========================================

    # A. Liste de base (Triée par Nom)
    students_list = Student.objects.filter(school=school).order_by('last_name', 'first_name')

    # B.  Récupérer la liste des Classes (Pour le filtre)
    # On utilise .exclude(class_name='') pour ne pas afficher les classes vides
    all_classes = Student.objects.filter(school=school).exclude(class_name='').values_list('class_name', flat=True).distinct().order_by('class_name')

    # C. Récupération des paramètres URL
    search_query = request.GET.get('q', '').strip()
    selected_class = request.GET.get('classe', '').strip()

    # D.  Application du filtre RECHERCHE
    if search_query:
        students_list = students_list.filter(
            Q(last_name__icontains=search_query) | 
            Q(first_name__icontains=search_query) | 
            Q(code_id__icontains=search_query) |
            Q(badge_number__icontains=search_query)
        )

    # E.  Application du filtre CLASSE
    if selected_class:
        students_list = students_list.filter(class_name=selected_class)

    # F. Pagination
    per_page = request.GET.get('per_page', '100')
    try:
        per_page = int(per_page)
        if per_page < 1: per_page = 100
    except:
        per_page = 100

    paginator = Paginator(students_list, per_page)
    page_number = request.GET.get('page')
    students = paginator.get_page(page_number)

    # ==========================================
    # 3. ENVOI AU TEMPLATE
    # ==========================================
    return render(request, 'manage_students.html', {
        'school': school, 
        'students': students,
        'per_page': per_page,
        
        #  Variables indispensables pour le filtre
        'all_classes': all_classes,
        'search_query': search_query,
        'selected_class': selected_class
    })
 
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def set_admin_badge(request):
    try:
        school = request.user.profile.school
        new_code = request.data.get('code')
        
        if not new_code:
            return Response({"error": "Code manquant"}, status=400)
            
        school.admin_badge_code = new_code
        school.save()
        
        return Response({"message": "Badge Admin mis à jour !"})
    except Exception as e:
        return Response({"error": str(e)}, status=400)



# --- 3. MOBILE SYNC API (Mise à jour Auto) ---
'''@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_school_config(request):
    try:
        student_id = request.GET.get('student_id')
        if not student_id: return Response({"error": "ID manquant"}, status=400)
        
        student = Student.objects.filter(code_id=student_id).first()
        if not student: 
            return Response({"error": "Non trouvé"}, status=404)
        #  1. نصايبو الرابط ديال التصويرة (نفس سمية Dossier و Fichier)
        school_folder = student.school.name.replace(" ", "_")
        # كنستعملو display_badge باش نجيبو الرقم لي مكتوب فالتصويرة
        display_badge = student.badge_number if student.badge_number else "---"
        img_filename = f"{student.last_name}_{student.first_name}_{display_badge}.png".replace(" ", "_")
        
        # الرابط الكامل (مثلا: http://ip:8000/media/qr_codes_ecoles/Ecole_X/Nom_Prenom_123.png)
        # ملاحظة: هادشي غايخدم إلا كانت التصويرة ديجا كاينة فالسيرفر
        qr_image_url = request.build_absolute_uri(f"{settings.MEDIA_URL}qr_codes_ecoles/{school_folder}/{img_filename}")
        
        return Response({
            "student_info": {
                "id": student.code_id,
                "first_name": student.first_name,
                "last_name": student.last_name,
                "class_name": student.class_name,
                "badge_number": student.badge_number,
                "qr_image_url": qr_image_url,
            },
            "school_config": {
                "admin_badge": student.school.admin_badge_code,
                "radius": student.school.radius,
                "lat": student.school.latitude,
                "lng": student.school.longitude,
                "pin": student.school.security_pin,
                "api_key": student.school.api_key,
                
            }
        })
    except Exception as e:
        return Response({"error": str(e)}, status=500)

        
'''
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_school_config(request):
    try:
        student_id = request.GET.get('student_id')
        if not student_id:
            return Response({"error": "ID manquant"}, status=400)

        student = Student.objects.filter(code_id=student_id).first()
        if not student:
            return Response({"error": "Non trouvé"}, status=404)

        # 1) QR image URL
        school_folder = (student.school.name or "").replace(" ", "_")
        display_badge = student.badge_number if student.badge_number else "---"
        img_filename = f"{student.last_name}_{student.first_name}_{display_badge}.png".replace(" ", "_")
        qr_image_url = request.build_absolute_uri(
            f"{settings.MEDIA_URL}qr_codes_ecoles/{school_folder}/{img_filename}"
        )

        school = student.school  # ✅ shortcut

        return Response({
            "student_info": {
                "id": student.code_id,
                "first_name": student.first_name,
                "last_name": student.last_name,
                "class_name": student.class_name,
                "badge_number": student.badge_number,
                "qr_image_url": qr_image_url,
            },
            "school_config": {
                # ✅✅✅ ADD THIS
                "school_id": school.id,          # <-- هادي اللي خاص Flutter
                "school_name": school.name,      # optional

                "admin_badge": school.admin_badge_code,
                "radius": school.radius,
                "lat": school.latitude,
                "lng": school.longitude,
                "pin": school.security_pin,
                "api_key": school.api_key,
            }
        })
    except Exception as e:
        return Response({"error": str(e)}, status=500)
     

@api_view(['GET'])
@permission_classes([AllowAny]) # ✅ أي واحد يقدر يشوفو (Public)
def public_school_scan(request, school_slug):
    try:
        # 1. البحث عن المدرسة بالاسم المختصر (Slug)
        # iexact كتعني مايهمش واش majuscule ولا minuscule
        school = School.objects.get(slug__iexact=school_slug)
        
        # 2. جلب الطلبات المعلقة (Pending) الخاصة بهاد المدرسة فقط
        pending_requests = PickupRequest.objects.filter(
            student__school=school, # ✅ الفلتر السحري
            is_completed=False
        ).order_by('-created_at') # الأحدث فالأعلى

        # 3. تجهيز البيانات JSON
        data = []
        for req in pending_requests:
            data.append({
                "id": req.id,
                "student_name": f"{req.student.last_name} {req.student.first_name}",
                "class_name": req.student.class_name,
                "badge_number": req.student.badge_number,
                "parent_name": f"{req.user.last_name} {req.user.first_name}",
                "time": req.created_at.strftime("%H:%M") # الوقت
            })

        return Response({
            "school": school.name,
            "count": len(data),
            "scans": data
        })

    except School.DoesNotExist:
        return Response({"error": "École introuvable (Vérifiez le slug)"}, status=404)
    except Exception as e:
        return Response({"error": str(e)}, status=500)
        

# Dans views.py
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def check_school_security(request):
    try:
        # On regarde l'école du parent connecté
        parent_school = request.user.profile.school
        
        return Response({
            "success": True,
            "school_config": {
                "requires_admin_scan": bool(parent_school.active_admin_scan),
                "admin_badge": parent_school.admin_badge_code if parent_school.active_admin_scan else None,
            }
        })
    except:
        # Si le parent n'a pas d'école, on considère que la sécurité est OFF
        return Response({"success": False, "requires_admin_scan": False})
        
        
@api_view(['GET'])
@permission_classes([AllowAny]) # أو IsAuthenticated إذا كان الوالد مسجلاً
def check_admin_scan_status(request):
    
    #  التصحيح: استعمل request.query_params أو request.GET
    api_key = request.query_params.get('api_key')

    if not api_key:
        return Response({'error': 'API Key manquant'}, status=400)

    try:
        # هنا كنقلبو على المدرسة بهاد الكود
        # ملاحظة: تأكد أن "13" هو بصح api_key فالجدول، ماشي id
        school = School.objects.get(api_key=api_key)
        
        return Response({
            'success': True, 
            'active': school.active_admin_scan
        })

    except School.DoesNotExist:
        return Response({'error': 'École introuvable avec ce code'}, status=404)
        
        

@api_view(['GET'])
@permission_classes([AllowAny]) # Tout le monde peut voir ça
def cgu_page(request):
    return render(request, 'Conditions_Utilisation.html')
    


@require_POST
def save_class_order(request):
    try:
        # 1. Njibo data li siftna (List d smiyat mrttbin)
        data = json.loads(request.body)
        new_order_list = data.get('order', [])
        
        # 2. Njibo l-madrassa dyal l-user courant
        school = request.user.profile.school
        
        # 3. N7awlo List l String (ex: ["MS", "GS"] -> "MS,GS")
        school.class_order = ",".join(new_order_list)
        school.save()
        
        return JsonResponse({'status': 'success'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})


def delete_class(request, class_id):
    # 1. N-vérifier wash l-user connect é
    if not request.user.is_authenticated:
        return redirect('login')
        
    school = request.user.profile.school
    
    # 2. Njibo l-classe w nt2kdo anaha dyal had l-madrassa
    classe_to_delete = get_object_or_404(SchoolClass, id=class_id, school=school)
    
    # 3. Nms7oha
    classe_to_delete.delete()
    
    # 4. Nraj3o l-user l manage_students m3a parametre bach l-modal tb9a m7loula
    base_url = reverse('manage_students')
    return redirect(f"{base_url}?show_config=true")
    



@require_http_methods(["GET"])
def school_ports(request, school_id):
    total = SchoolPort.objects.filter(school_id=school_id).count()
    active = SchoolPort.objects.filter(school_id=school_id, is_active=True).count()

    ports = SchoolPort.objects.filter(school_id=school_id, is_active=True).order_by("order")
    data = [{"key": p.key, "label": p.label, "is_active": p.is_active} for p in ports]

    return JsonResponse({"school_id": school_id, "total": total, "active": active, "ports": data})



@csrf_exempt
@require_http_methods(["POST"])
def rename_port_api(request, school_id):
    body = json.loads(request.body.decode("utf-8"))
    key = body.get("key")
    label = body.get("label")

    if not key or not label:
        return JsonResponse(
            {"ok": False, "error": "key/label required"},
            status=400 
        )

    p = SchoolPort.objects.get(school_id=school_id, key=key)
    p.label = label.strip()[:50] 
    p.save(update_fields=["label"])

    return JsonResponse({"ok": True, "key": p.key, "label": p.label})


def sanitize_filename(name):
    clean_name = re.sub(r'[^a-zA-Z0-9\s_-]', '', name)
    clean_name = clean_name.strip().replace(' ', '_')
    return clean_name

def pickup_historybkp(request):
    # تحقق من المدرسة
    if not hasattr(request.user, 'profile') or not request.user.profile.school:
        return JsonResponse({'error': 'Aucune école assignée'}, status=403)

    current_school = request.user.profile.school

    # params
    date_from_str = request.GET.get('date_from')
    date_to_str = request.GET.get('date_to')
    export_excel = request.GET.get('export') == 'true'

    filter_start = datetime.strptime(date_from_str, '%Y-%m-%d').date() if date_from_str else None
    filter_end   = datetime.strptime(date_to_str, '%Y-%m-%d').date() if date_to_str else None

    # Query DB فقط
    qs = PickupRequest.objects.filter(student__school=current_school)\
        .select_related('student')\
        .order_by('-created_at')

    if filter_start:
        qs = qs.filter(created_at__date__gte=filter_start)
    if filter_end:
        qs = qs.filter(created_at__date__lte=filter_end)

    # Excel
    if export_excel:
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f"Historique_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sorties"

        headers = ["Date Heure", "Élève", "Classe", "Badge", "Porte", "État"]
        ws.append(headers)

        for p in qs:
            dt = p.created_at
            if timezone.is_aware(dt):
                dt = timezone.localtime(dt)

            ws.append([
                dt.strftime('%d-%m-%Y %H:%M:%S'),
                f"{p.student.first_name} {p.student.last_name}",
                p.student.class_name or "",
                p.student.badge_number or "-",
                p.porte_label or "",
                "Completed" if p.is_completed else "Pending",
            ])

        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 15

        wb.save(response)
        return response

    # JSON response
    data = []
    for p in qs:
        dt = p.created_at
        if timezone.is_aware(dt):
            dt = timezone.localtime(dt)

        data.append({
            'date': dt.strftime('%d-%m-%Y %H:%M:%S'),
            'student': f"{p.student.first_name} {p.student.last_name}",
            'classe': p.student.class_name or "",
            'porte': p.porte_label or "",
            'status': "Completed" if p.is_completed else "Pending",
            'badge_number': p.student.badge_number or "-"
        })

    return JsonResponse({'data': data})


def pickup_historybkp2(request):
    # تحقق من المدرسة
    if not hasattr(request.user, 'profile') or not request.user.profile.school:
        return JsonResponse({'error': 'Aucune école assignée'}, status=403)

    current_school = request.user.profile.school

    # params
    date_from_str = request.GET.get('date_from')
    date_to_str = request.GET.get('date_to')
    export_excel = request.GET.get('export') == 'true'

    filter_start = datetime.strptime(date_from_str, '%Y-%m-%d').date() if date_from_str else None
    filter_end   = datetime.strptime(date_to_str, '%Y-%m-%d').date() if date_to_str else None

    # Query DB باستعمال موديل Backup
    # كنقلبو مباشرة بالسمية ديال المدرسة ومكنحتاجوش select_related
    qs = PickupRequestBackup.objects.filter(school_name=current_school.name)\
        .order_by('-created_at_original')

    # الفيلتر بالتاريخ الأصلي
    if filter_start:
        qs = qs.filter(created_at_original__date__gte=filter_start)
    if filter_end:
        qs = qs.filter(created_at_original__date__lte=filter_end)

    # Excel Export
    if export_excel:
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f"Historique_Backup_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sorties"

        headers = ["Date Heure", "Élève", "Classe", "Badge", "Porte", "État"]
        ws.append(headers)

        for p in qs:
            dt = p.created_at_original # التاريخ من الموديل الجديد
            if timezone.is_aware(dt):
                dt = timezone.localtime(dt)

            ws.append([
                dt.strftime('%d-%m-%Y %H:%M:%S'),
                p.student_name or "",      # كنجبدو السمية مباشرة
                p.class_name or "",        # القسم مباشرة
                p.badge_number or "-",
                p.porte_label or "",
                "Completed" if p.is_completed else "Pending",
            ])

        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 15

        wb.save(response)
        return response

    # JSON response
    data = []
    for p in qs:
        dt = p.created_at_original # التاريخ من الموديل الجديد
        if timezone.is_aware(dt):
            dt = timezone.localtime(dt)

        data.append({
            'date': dt.strftime('%d-%m-%Y %H:%M:%S'),
            'student': p.student_name or "", # السمية مباشرة
            'classe': p.class_name or "",    # القسم مباشرة
            'porte': p.porte_label or "",
            'status': "Completed" if p.is_completed else "Pending",
            'badge_number': p.badge_number or "-"
        })

    return JsonResponse({'data': data})


@api_view(['POST'])
@permission_classes([AllowAny])
def custom_login(request):
    username = request.data.get('username')
    password = request.data.get('password')

    # تنظيف اليوزرنيم من أي فراغات زايدة تقدر تجي من التليفون أو الكيبورد
    if username:
        username = username.strip()

    print(f"LOGIN ATTEMPT: '{username}'")  

    # ⚠️ التعديل الأهم: زدنا `request` هنا باش django-axes يخدم بشكل صحيح
    user = authenticate(request, username=username, password=password)

    if user is not None:
        print("User found in DB") 
        try:
            profile = user.profile
            print(f"Profile found: Parent={profile.is_parent_account}") 
        except UserProfile.DoesNotExist: # تأكد أن UserProfile مستوردة الفوق
            print("ERROR: Profile NOT FOUND") 
            return Response({"error": "Admin Error: User Profile Missing"}, status=400)

        token, _ = Token.objects.get_or_create(user=user)

        return Response({
            'token': token.key,
            'is_parent': profile.is_parent_account,
            'school_id': profile.school.id,
            'school_name': profile.school.name,
            'school_lat': profile.school.latitude,
            'school_lng': profile.school.longitude,
            'radius': profile.school.radius,
            'admin_badge': profile.school.admin_badge_code,
            'security_pin': profile.school.security_pin,
            'school_api_key': profile.school.api_key,
        })
    
    else:
        # إلا بقى كيعطي خطأ، هاد الـ print غتبين ليك واش المشكل في الباسورد أو اليوزرنيم
        print(f"FAILED AUTH FOR: '{username}'") 
        return Response({"error": "Identifiants incorrects"}, status=400)
        
    
 



def pickup_history(request):
    # تحقق من المدرسة
    if not hasattr(request.user, 'profile') or not request.user.profile.school:
        return JsonResponse({'error': 'Aucune école assignée'}, status=403)

    current_school = request.user.profile.school

    # params
    date_from_str = request.GET.get('date_from')
    date_to_str = request.GET.get('date_to')
    export_excel = request.GET.get('export') == 'true'

    filter_start = datetime.strptime(date_from_str, '%Y-%m-%d').date() if date_from_str else None
    filter_end   = datetime.strptime(date_to_str, '%Y-%m-%d').date() if date_to_str else None

    # Query DB باستعمال موديل Backup
    qs = PickupRequestBackup.objects.filter(school_name=current_school.name)\
        .order_by('-created_at_original')

    # الفيلتر بالتاريخ الأصلي
    if filter_start:
        qs = qs.filter(created_at_original__date__gte=filter_start)
    if filter_end:
        qs = qs.filter(created_at_original__date__lte=filter_end)

    # ---------------------------------------------------------
    # حجب التكرار (Deduplication) ديال أجزاء الميلي ثانية
    # ---------------------------------------------------------
    unique_records = []
    last_scan_time = {}  # ديكسيونير باش نعقلو على آخر مرة تسكانا فيها البادج

    for p in qs:
        badge = p.badge_number
        dt = p.created_at_original
        
        # إذا كان البادج كاين، نفيليتريوه
        if badge and badge != "-":
            if badge in last_scan_time:
                # نحسبو الفرق بالثواني بين هاد التسجيلة والتسجيلة لي قبلها
                time_diff = abs((last_scan_time[badge] - dt).total_seconds())
                if time_diff < 5:  # 5 ثواني (تقدر تبدلها لـ 2 أو 3 يلا بغيتي)
                    continue  # تجاهل هاد التسجيلة (ماتزيدهاش)
            
            # تحديث وقت آخر سكان لهاد البادج
            last_scan_time[badge] = dt
            
        unique_records.append(p)

    # دابا غنخدمو بـ unique_records بلاصة qs
    # Excel Export
    if export_excel:
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f"Historique_Backup_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sorties"

        headers = ["Date Heure", "Élève", "Classe", "Badge", "Porte", "État"]
        ws.append(headers)

        for p in unique_records:  # <-- استعملنا الليستا المفلترة
            dt = p.created_at_original
            if timezone.is_aware(dt):
                dt = timezone.localtime(dt)

            ws.append([
                dt.strftime('%d-%m-%Y %H:%M:%S'),
                p.student_name or "",      
                p.class_name or "",        
                p.badge_number or "-",
                p.porte_label or "",
                "Completed" if p.is_completed else "Pending",
            ])

        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 15

        wb.save(response)
        return response

    # JSON response
    data = []
    for p in unique_records:  # <-- استعملنا الليستا المفلترة
        dt = p.created_at_original
        if timezone.is_aware(dt):
            dt = timezone.localtime(dt)

        data.append({
            'date': dt.strftime('%d-%m-%Y %H:%M:%S'),
            'student': p.student_name or "", 
            'classe': p.class_name or "",    
            'porte': p.porte_label or "",
            'status': "Completed" if p.is_completed else "Pending",
            'badge_number': p.badge_number or "-"
        })

    return JsonResponse({'data': data})



def school_classes_list(request):
    # كنجيبو المدرسة ديال الحساب اللي فاتح
    user_school = request.user.profile.school 
    
    # كنجيبو أسماء الأقسام "الفريدة" من جدول AllStudents
    # استعملنا .distinct() باش ميتعاودش لينا نفس القسم بزاف دالمرات
    classes = AllStudents.objects.filter(school=user_school).values('class_name').distinct()
    
    return render(request, 'school_classes_list.html', {
        'classes': classes,
        'school': user_school
    })


def manage_attendance(request, api_key, class_name):
    user_school = request.user.profile.school
    today = now().date()
    students = AllStudents.objects.filter(school=user_school, class_name=class_name)
    decoded_class_name = urllib.parse.unquote(class_name)

    if request.method == 'POST':
        for student in students:
            status = request.POST.get(f'status_{student.id}')
            if status:
                minutes = request.POST.get(f'minutes_{student.id}', 0) if status == 'LATE' else 0
                try:
                    minutes = int(minutes) if minutes else 0
                except ValueError:
                    minutes = 0
                # كنسجلوا سطر جديد لكل حصة
                # هكذا الحصة 1 كتبقى ABSENT والحصة 3 تقدر تولي PRESENT
                StudentPresence.objects.create(
                    student=student,
                    status=status,
                    school=user_school,
                    retard_minutes=minutes
                    # الوقت كيتسجل تلقائياً بالدقيقة والثانية
                )
        return redirect('school_classes_list')

    # لعرض "آخر حالة" فقط في صفحة التسجيل (الأقفال)
    attendance_map = {}
    for student in students:
        last_p = StudentPresence.objects.filter(student=student, date__date=today).last()
        if last_p:
            attendance_map[student.id] = last_p.status

    context = {
        'students': students,
        'class_name': decoded_class_name,
        'attendance_map': attendance_map,
        'today_date': today,
    }
    return render(request, 'attendance_form.html', context)




@login_required
def admin_presence_dashboard(request):
    user_school = request.user.profile.school
    
    # 1. التقاط تواريخ الفلتر
    start_date_str = request.GET.get('start')
    end_date_str = request.GET.get('end')
    
    if start_date_str and end_date_str:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        except ValueError:
            start_date = timezone.now().date()
            end_date = timezone.now().date()
    else:
        start_date = timezone.now().date()
        end_date = timezone.now().date()

    # تحديد الوقت من 00:00 إلى 23:59
    start_datetime = datetime.combine(start_date, time.min)
    end_datetime = datetime.combine(end_date, time.max)
    
    if timezone.is_aware(timezone.now()):
        start_datetime = timezone.make_aware(start_datetime)
        end_datetime = timezone.make_aware(end_datetime)

    # 2. جلب جميع السجلات للفترة المحددة
    all_records = StudentPresence.objects.filter(
        school=user_school,
        date__range=(start_datetime, end_datetime)
    ).select_related('student')

    # 3. تصدير XLSX (اختياري)
    if request.GET.get('export') == 'xlsx':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Rapport Absences"

        # 1. تحضير ستايلات الديزاين الاحترافي
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        alt_row_fill = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")
        center_aligned = Alignment(horizontal="center", vertical="center")
        left_aligned = Alignment(horizontal="left", vertical="center")
        thin_border = Border(
            left=Side(style='thin', color='D4D4D4'),
            right=Side(style='thin', color='D4D4D4'),
            top=Side(style='thin', color='D4D4D4'),
            bottom=Side(style='thin', color='D4D4D4')
        )

        # 2. إضافة العناوين
        headers = ['Date', 'Heure', 'Élève', 'Classe', 'Statut', 'Retard (Min)']
        ws.append(headers)

        for col_num, cell in enumerate(ws[1], 1):
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_aligned
            cell.border = thin_border

        # 3. تصفية البيانات (هنا فين حيدنا الحاضرين)
        # هاد السطر كيجمع غير التلاميذ اللي الحالة ديالهم ماشي PRESENT
        filtered_records = [r for r in all_records if r.status != 'PRESENT']

        # إضافة البيانات اللي بقات (الغياب والتأخر فقط)
        for row_num, record in enumerate(filtered_records, start=2):
            minutes = record.retard_minutes if record.status == 'LATE' and record.retard_minutes else '-'
            
            row_data = [
                record.date.strftime('%Y-%m-%d'),
                record.date.strftime('%H:%M'),
                f"{record.student.first_name} {record.student.last_name}",
                record.student.class_name,
                record.get_status_display(),
                minutes
            ]
            ws.append(row_data)

            # تطبيق الديزاين على أسطر البيانات
            for col_num, cell in enumerate(ws[row_num], 1):
                cell.border = thin_border
                if col_num == 3:
                    cell.alignment = left_aligned
                else:
                    cell.alignment = center_aligned
                
                if row_num % 2 == 0:
                    cell.fill = alt_row_fill

        # 4. تعديل عرض الأعمدة أوتوماتيكياً
        column_widths = [15, 12, 35, 15, 15, 15]
        for i, column_width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = column_width

        # 5. إرسال الملف
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="Rapport_Absences_Retards.xlsx"'
        wb.save(response)
        return response

    # 4. بناء الإحصائيات لجميع الأقسام (حتى اللي ما فيهومش غياب)
    # غنجيبو كاع الأقسام اللي مسجلين فالمدرسة
    all_classes = SchoolClass.objects.filter(school=user_school).order_by('name')
    
    stats = []
    js_classes = []
    js_absents = []
    js_lates = []

    for cls in all_classes:
        # جلب السجلات الخاصة بهاد القسم فقط
        class_records = all_records.filter(student__class_name=cls.name)
        
        absents_count = class_records.filter(status='ABSENT').count()
        lates_count = class_records.filter(status='LATE').count()
        # جلب عدد التلاميذ المسجلين فهاد القسم
        total_students = AllStudents.objects.filter(school=user_school, class_name=cls.name).count()
        
        stats.append({
            'student__class_name': cls.name,
            'absents': absents_count,
            'lates': lates_count,
            'total_students': total_students
        })
        
        # بيانات المبيان
        js_classes.append(cls.name)
        js_absents.append(absents_count)
        js_lates.append(lates_count)

    # 5. منطق الجورنال (Journal)
    sessions_dict = {}
    issues_only = all_records.exclude(status='PRESENT')
    
    for issue in issues_only:
        date_str = issue.date.strftime('%Y-%m-%d')
        time_str = issue.date.strftime('%H:%M')
        key = (issue.student.class_name, date_str, time_str)
        
        if key not in sessions_dict:
            sessions_dict[key] = {
                'type': 'ISSUE',
                'class_name': issue.student.class_name,
                'date_display': date_str,
                'time': time_str,
                'timestamp': issue.date,
                'students': []
            }
        sessions_dict[key]['students'].append(issue)

    # إضافة الحصص السليمة
    for r in all_records:
        date_str = r.date.strftime('%Y-%m-%d')
        time_str = r.date.strftime('%H:%M')
        key = (r.student.class_name, date_str, time_str)
        if key not in sessions_dict:
            sessions_dict[key] = {
                'type': 'PERFECT',
                'class_name': r.student.class_name,
                'date_display': date_str,
                'time': time_str,
                'timestamp': r.date,
                'students': []
            }

    combined_journal = list(sessions_dict.values())
    combined_journal.sort(key=lambda x: x['timestamp'], reverse=True)

    context = {
        'stats': stats,
        'combined_journal': combined_journal,
        'start_date': start_date.strftime('%Y-%m-%d'),
        'end_date': end_date.strftime('%Y-%m-%d'),
        'total_absent': all_records.filter(status='ABSENT').count(),
        'total_late': all_records.filter(status='LATE').count(),
        'js_classes': json.dumps(js_classes),
        'js_absents': json.dumps(js_absents),
        'js_lates': json.dumps(js_lates),
    }
    
    return render(request, 'admin_presence_dashboard.html', context)
    
    
    

@login_required
@require_POST
def delete_school_class(request, class_name):
    user_school = request.user.profile.school
    # مسح القسم من جدول الأقسام للمدرسة الحالية
    SchoolClass.objects.filter(school=user_school, name=class_name).delete()
    # مسح التلاميذ اللي مسجلين فهاد القسم للمدرسة الحالية
    AllStudents.objects.filter(school=user_school, class_name=class_name).delete()
    
    return JsonResponse({'status': 'success'})
    
    
    
@login_required  
def get_students_by_class(request, class_name):
    user_school = request.user.profile.school
    students = AllStudents.objects.filter(school=user_school, class_name=class_name)
    
    data = []
    for s in students:
        # حساب مجموع دقائق التأخر لهذا التلميذ فقط
        total_minutes = StudentPresence.objects.filter(
            student=s, 
            status='LATE'
        ).aggregate(Sum('retard_minutes'))['retard_minutes__sum'] or 0
        
        data.append({
            'id': s.id,
            'first_name': s.first_name,
            'last_name': s.last_name,
            'total_retard': total_minutes  # صيفطنا المجموع هنا
        })
    return JsonResponse(data, safe=False)

# دالة الحذف (تأكد أنها موجودة)
def delete_presence(request, presence_id):
    presence = get_object_or_404(StudentPresence, id=presence_id, school=request.user.profile.school)
    presence.delete()
    return redirect('admin_presence_dashboard')
    
def download_template(request):
    # إنشاء ملف Excel جديد
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Liste Eleves"
    
    # كتابة العناوين
    headers = ['NOM', 'PRENOM', 'CLASSE']
    sheet.append(headers)
    
    # إضافة مثال (اختياري)
    sheet.append(['ALAMI', 'AHMED', 'CE6-A'])
    
    # إرسال الملف للمتصفح
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="modele_eleves.xlsx"'
    wb.save(response)
    return response    
    
    
def upload_all_students_excel(request):
    print(f"DEBUG: Method is {request.method}")
    
    # تحديد اسم القالب (Template) ديال هاد الصفحة
    # (بدل 'your_template_name.html' بالسمية ديال الـ html ديالك)
    template_name = 'upload_excel.html' 

    if request.method == "POST":
        try:
            # 1. جلب الملف
            excel_file = request.FILES.get('file')
            print(f"DEBUG: File received: {excel_file}")

            if not excel_file:
                messages.error(request, "الملف لم يصل للسيرفر! المرجو اختيار ملف.")
                return render(request, template_name) # البقاء في نفس الصفحة

            # 2. هاد السطر دخلناه لوسط try باش يلا كان مشكل فـ profile يتشد
            user_school = request.user.profile.school
            
            # 3. قراءة ملف الإكسيل
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active
            
            count = 0
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row[0] or not row[1] or not row[2]:
                    continue
                
                nom = str(row[0]).strip()
                prenom = str(row[1]).strip()
                classe = str(row[2]).strip()

                # إنشاء القسم والتلميذ
                SchoolClass.objects.get_or_create(school=user_school, name=classe)
                AllStudents.objects.get_or_create(
                    school=user_school,
                    first_name=prenom,
                    last_name=nom,
                    class_name=classe
                )
                count += 1
            
            print(f"DEBUG: Success! Added {count} students.")
            messages.success(request, f"تم بنجاح إضافة {count} تلميذ!")
            return redirect('school_classes_list') 

        except Exception as e:
            # 4. أي خطأ وقع (سواء فالاكسيل ولا فالداتا) غيطيح هنا
            
            messages.error(request, f" Vérifiez votre fichier svp Merci")
            return render(request, template_name) # البقاء في نفس الصفحة مع إظهار الخطأ

    # 5. هاد Return ضرورية باش الصفحة تفتح فاش اليوزر كيدخل ليها أول مرة (GET request)
    return render(request, template_name)   
   
    
@login_required
def school_portal(request):
    return render(request, 'portal.html')
    
def clear_all_students(request):
    if request.method == "POST":
        # كنحددو المدرسة ديال الشخص اللي داخل دابا (Admin المتصل)
        user_school = request.user.profile.school
        
        # كنمسحو غير التلاميذ اللي تابعين لهاد المدرسة بالضبط
        AllStudents.objects.filter(school=user_school).delete()
        
        messages.success(request, f"تم حذف قائمة تلاميذ مدرسة {user_school.name} بالكامل.")
        
    return redirect('admin_presence_dashboard')
    
@login_required
def save_student_ajax(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        student_id = data.get('id')
        first_name = data.get('first_name')
        last_name = data.get('last_name')
        class_name = data.get('class_name')
        user_school = request.user.profile.school

        if student_id:  # إذا كاين ID يعني هادا "تعديل" (Modifier)
            student = AllStudents.objects.get(id=student_id, school=user_school)
            student.first_name = first_name
            student.last_name = last_name
            student.save()
            return JsonResponse({'status': 'success', 'message': 'Élève modifié avec succès'})
        else:  # إذا ماكاينش ID يعني هادا "إضافة" (Ajouter)
            AllStudents.objects.create(
                school=user_school,
                first_name=first_name,
                last_name=last_name,
                class_name=class_name
            )
            return JsonResponse({'status': 'success', 'message': 'Élève ajouté avec succès'})
    return JsonResponse({'status': 'error'}, status=400)

@login_required
def delete_student_ajax(request, student_id):
    if request.method == 'DELETE':
        student = AllStudents.objects.get(id=student_id, school=request.user.profile.school)
        student.delete()
        return JsonResponse({'status': 'success'})
    return JsonResponse({'status': 'error'}, status=400)
    
    
def export_all_students(request):
    user_school = request.user.profile.school
    students = AllStudents.objects.filter(school=user_school).order_by('class_name', 'last_name')
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tous les élèves"
    
    # العناوين
    ws.append(['Nom', 'Prénom', 'Classe'])
    
    for s in students:
        ws.append([s.last_name, s.first_name, s.class_name])
        
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="tous_les_eleves.xlsx"'
    wb.save(response)
    return response

# 2. إخراج قسم معين
def export_class_students(request, class_name):
    user_school = request.user.profile.school
    students = AllStudents.objects.filter(school=user_school, class_name=class_name).order_by('last_name')
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Classe {class_name}"
    
    ws.append(['Nom', 'Prénom', 'Classe'])
    
    for s in students: 
        ws.append([s.last_name, s.first_name, s.class_name])
        
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="classe_{class_name}.xlsx"'
    wb.save(response)
    return response 
    


def search_student(request):
    query = request.GET.get('q', '')
    user_school = request.user.profile.school
    students = AllStudents.objects.filter(
        school=user_school
    ).filter(
        Q(first_name__icontains=query) | Q(last_name__icontains=query)
    )[:10] # أهم 10 نتائج
    
    data = [{'id': s.id, 'name': f"{s.last_name} {s.first_name}", 'class': s.class_name} for s in students]
    return JsonResponse(data, safe=False)


    

    
def student_history(request, student_id):
    # كنجيبو كاع الحالات من غير الحاضرين
    presences = StudentPresence.objects.filter(student_id=student_id).exclude(status='PRESENT').order_by('-date')
    
    data = []
    for p in presences:
        data.append({
            'date': p.date.strftime('%d/%m/%Y %H:%M'),
            'status': p.get_status_display(), # النص (Absent / Retard)
            'status_code': p.status,          # الكود (ABSENT / LATE)
            'retard_minutes': p.retard_minutes or 0 ,# الدقائق
            'class_name': p.student.class_name
        })
    return JsonResponse(data, safe=False)